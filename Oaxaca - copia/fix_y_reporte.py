#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fix de JSONs de Oaxaca + Generación de Reporte Excel
- Agrega campos v8 faltantes: metodo_extraccion, calidad_porcentaje, calidad_detalle
- Intenta extraer fecha_ultima_reforma del contenido cuando está vacío
- Detecta y reporta URLs duplicadas
- Genera reporte Excel completo
"""

import os
import re
import json
import glob
from datetime import datetime

BASE = "/mnt/c/Users/nayos/Documents/mi-Claude/Oaxaca"
TEXTOS_DIR = os.path.join(BASE, "Textos de Oaxaca")

SECCIONES = ["Estatal", "Federal", "Marco Normativo", "Municipal"]


# ============================================================
# CALIDAD (misma fórmula v8 que los otros scrapers)
# ============================================================

def calcular_calidad(contenido, numero_paginas, metodo):
    if not contenido or not contenido.strip():
        return {"calidad_porcentaje": 0, "calidad_detalle": "Sin contenido"}
    chars = len(contenido)
    paginas = max(numero_paginas, 1)
    score_densidad = min((chars / paginas) / 2000, 1.0)
    alfanum = sum(1 for c in contenido if c.isalnum() or c in ' \n\t.,;:()áéíóúñüÁÉÍÓÚÑÜ"\'—–-')
    ratio_alfanum = alfanum / max(chars, 1)
    lineas = contenido.split('\n')
    lineas_con_texto = sum(1 for l in lineas if len(l.strip()) > 5)
    ratio_lineas = lineas_con_texto / max(len(lineas), 1)
    palabras = contenido.split()
    palabras_cortas = sum(1 for p in palabras if len(p) == 1 and p.lower() not in 'aeiouy')
    ratio_basura = 1.0 - min(palabras_cortas / max(len(palabras), 1) * 5, 0.5)
    calidad = score_densidad * 0.30 + ratio_alfanum * 0.25 + ratio_lineas * 0.20 + ratio_basura * 0.25
    porcentaje = round(min(calidad * 100, 100))
    if porcentaje >= 85:   detalle = "Excelente"
    elif porcentaje >= 70: detalle = "Buena"
    elif porcentaje >= 50: detalle = "Aceptable"
    elif porcentaje >= 30: detalle = "Baja"
    else:                  detalle = "Muy baja"
    return {"calidad_porcentaje": porcentaje, "calidad_detalle": detalle}


# ============================================================
# EXTRACCIÓN DE FECHA DE REFORMA DESDE CONTENIDO
# ============================================================

def extraer_fecha_reforma_contenido(titulo, contenido):
    meses = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
    }
    fuentes = []
    if contenido:
        fuentes.append(contenido[:15000].lower())
    if titulo:
        fuentes.append(titulo.lower())

    for fuente in fuentes:
        # "Última reforma" explícita
        patron_ult = r'[uú]ltima\s+reforma\s*:?\s*[^\n]{0,120}?(\d{1,2})[°º]?\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de(?:l)?\s+(\d{4})'
        match = re.search(patron_ult, fuente)
        if match:
            return f"{match.group(1).zfill(2)}/{meses[match.group(2)]}/{match.group(3)}"

    if not contenido:
        return ""
    texto_lower = contenido.lower()

    # Buscar "reforma" con fecha
    patron_reforma = r'(?:[uú]ltima\s+)?reform[aáo][^.]{0,80}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de(?:l)?\s+(\d{4})'
    matches = re.findall(patron_reforma, texto_lower[:20000])
    if matches:
        fechas = []
        for m in matches:
            try:
                dia, anio = int(m[0]), int(m[2])
                mes_num = int(meses.get(m[1], 1))
                if 1900 < anio < 2100 and 1 <= dia <= 31:
                    fechas.append((anio, mes_num, dia, f"{str(dia).zfill(2)}/{meses[m[1]]}/{m[2]}"))
            except (ValueError, KeyError):
                continue
        if fechas:
            fechas.sort(reverse=True)
            return fechas[0][3]

    return ""


# ============================================================
# DETECTAR MÉTODO DE EXTRACCIÓN DESDE CONTENIDO
# ============================================================

def detectar_metodo(contenido, es_escaneado):
    """Infiere el método de extracción a partir del contenido."""
    if not contenido or len(contenido.strip()) < 50:
        return "ninguno"
    if "[Pág" in contenido and "OCR]" in contenido:
        # Tiene marcas de OCR
        if es_escaneado:
            return "pdfplumber+paddle"
        else:
            return "pdfplumber+paddle"
    if es_escaneado:
        return "pdfplumber+paddle"
    return "pdfplumber"


# ============================================================
# FIX PRINCIPAL
# ============================================================

def fix_jsons():
    print("=" * 65)
    print("  FIX JSONs OAXACA — Agregar campos v8 faltantes")
    print("=" * 65)

    total_fixed = 0
    total_fecha_fixed = 0
    total_files = 0

    for seccion in SECCIONES:
        json_dir = os.path.join(TEXTOS_DIR, seccion)
        if not os.path.exists(json_dir):
            continue

        archivos = sorted(glob.glob(os.path.join(json_dir, "*.json")))
        fixed_sec = 0
        fecha_fixed_sec = 0

        print(f"\n  {seccion}: {len(archivos)} archivos")

        for ruta in archivos:
            total_files += 1
            try:
                with open(ruta, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            except Exception:
                continue

            modificado = False

            contenido = data.get("contenido", "")
            titulo = data.get("titulo", "")
            num_pags = data.get("numero_paginas", 0) or 0
            es_escaneado = data.get("es_escaneado", False)

            # 1. Agregar metodo_extraccion si falta
            if "metodo_extraccion" not in data:
                data["metodo_extraccion"] = detectar_metodo(contenido, es_escaneado)
                modificado = True

            # 2. Agregar calidad_porcentaje y calidad_detalle si faltan
            if "calidad_porcentaje" not in data or "calidad_detalle" not in data:
                cal = calcular_calidad(contenido, num_pags, data.get("metodo_extraccion", ""))
                data["calidad_porcentaje"] = cal["calidad_porcentaje"]
                data["calidad_detalle"] = cal["calidad_detalle"]
                modificado = True

            # 3. Intentar llenar fecha_ultima_reforma si está vacía
            if not data.get("fecha_ultima_reforma"):
                fecha = extraer_fecha_reforma_contenido(titulo, contenido)
                if fecha:
                    data["fecha_ultima_reforma"] = fecha
                    fecha_fixed_sec += 1
                    modificado = True

            if modificado:
                with open(ruta, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                fixed_sec += 1

        total_fixed += fixed_sec
        total_fecha_fixed += fecha_fixed_sec
        print(f"    Campos v8 agregados: {fixed_sec}")
        print(f"    Fechas reforma recuperadas: {fecha_fixed_sec}")

    print(f"\n{'='*65}")
    print(f"  TOTAL: {total_fixed}/{total_files} JSONs actualizados")
    print(f"  Fechas reforma recuperadas: {total_fecha_fixed}")
    print(f"{'='*65}")
    return total_fixed, total_fecha_fixed


# ============================================================
# REPORTE EXCEL
# ============================================================

def generar_reporte():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl no instalado. pip install openpyxl")
        return

    print("\n" + "=" * 65)
    print("  GENERANDO REPORTE EXCEL — OAXACA")
    print("=" * 65)

    wb = Workbook()
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    fill_alt = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    fill_excelente = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fill_buena = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    fill_baja = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9E2F3'), right=Side(style='thin', color='D9E2F3'),
        top=Side(style='thin', color='D9E2F3'), bottom=Side(style='thin', color='D9E2F3'),
    )

    # ── HOJA 1: Resumen General ──
    ws = wb.active
    ws.title = "Resumen General"

    ws.merge_cells('A1:I1')
    ws['A1'] = "Reporte de Legislación — Congreso del Estado de Oaxaca"
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Sección', 'JSONs', 'PDFs', 'Con Texto', 'Sin Texto', 'Calidad Prom.',
               'URLs Dup.', 'Con Fecha Reforma', 'Páginas Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row = 5
    pdfs_base = os.path.join(BASE, "Leyes de Oaxaca")

    for seccion in SECCIONES:
        json_dir = os.path.join(TEXTOS_DIR, seccion)
        pdf_dir = os.path.join(pdfs_base, seccion)

        n_json = n_pdf = con_texto = sin_texto = pags = con_fecha = 0
        calidades = []
        urls = set()
        dups = 0

        if os.path.exists(json_dir):
            for jf in os.listdir(json_dir):
                if not jf.endswith('.json'):
                    continue
                n_json += 1
                try:
                    with open(os.path.join(json_dir, jf), 'r', encoding='utf-8') as f:
                        d = json.load(f)
                    c = d.get('contenido', '')
                    u = d.get('url', '')
                    calidades.append(d.get('calidad_porcentaje', 0))
                    pags += d.get('numero_paginas', 0) or 0
                    if c and len(c) > 100:
                        con_texto += 1
                    else:
                        sin_texto += 1
                    if d.get('fecha_ultima_reforma'):
                        con_fecha += 1
                    if u in urls:
                        dups += 1
                    urls.add(u)
                except:
                    pass

        if os.path.exists(pdf_dir):
            n_pdf = len([f for f in os.listdir(pdf_dir) if f.lower().endswith('.pdf')])

        cal = sum(calidades) / max(len(calidades), 1)

        ws.cell(row=row, column=1, value=seccion).border = thin_border
        for col, val in enumerate([n_json, n_pdf, con_texto, sin_texto, f"{cal:.0f}%", dups, con_fecha, pags], 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Color calidad
        cell_cal = ws.cell(row=row, column=6)
        if cal >= 85:
            cell_cal.fill = fill_excelente
        elif cal >= 70:
            cell_cal.fill = fill_buena
        elif cal > 0:
            cell_cal.fill = fill_baja

        row += 1

    ws.column_dimensions['A'].width = 22
    for c in range(2, 10):
        ws.column_dimensions[get_column_letter(c)].width = 15
    ws.freeze_panes = 'A5'

    # ── HOJAS POR SECCIÓN ──
    det_headers = ['#', 'Título', 'Ordenamiento', 'Status', 'Materia', 'Método',
                   'Páginas', 'Calidad %', 'Calidad', 'Fecha Pub.', 'Fecha Reforma',
                   'Escaneado', 'Caracteres', 'URL']

    for seccion in SECCIONES:
        json_dir = os.path.join(TEXTOS_DIR, seccion)
        if not os.path.exists(json_dir):
            continue

        ws_sec = wb.create_sheet(title=seccion[:31])
        for col, h in enumerate(det_headers, 1):
            cell = ws_sec.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        r = 2
        for jf in sorted(os.listdir(json_dir)):
            if not jf.endswith('.json'):
                continue
            try:
                with open(os.path.join(json_dir, jf), 'r', encoding='utf-8') as f:
                    d = json.load(f)
            except:
                continue

            num_match = re.match(r'^(\d+)', jf)
            materias = d.get('materia', [])
            if isinstance(materias, list):
                materias = ', '.join(materias)

            ws_sec.cell(row=r, column=1, value=int(num_match.group(1)) if num_match else r - 1).border = thin_border
            ws_sec.cell(row=r, column=2, value=d.get('titulo', '')[:200]).border = thin_border
            ws_sec.cell(row=r, column=3, value=d.get('ordenamiento', '')).border = thin_border
            ws_sec.cell(row=r, column=4, value=d.get('status', '')).border = thin_border
            ws_sec.cell(row=r, column=5, value=materias).border = thin_border
            ws_sec.cell(row=r, column=6, value=d.get('metodo_extraccion', '')).border = thin_border
            ws_sec.cell(row=r, column=7, value=d.get('numero_paginas', 0)).border = thin_border

            cal = d.get('calidad_porcentaje', 0)
            cell_cal = ws_sec.cell(row=r, column=8, value=cal)
            cell_cal.border = thin_border
            if cal >= 85:
                cell_cal.fill = fill_excelente
            elif cal >= 70:
                cell_cal.fill = fill_buena
            elif cal > 0:
                cell_cal.fill = fill_baja

            ws_sec.cell(row=r, column=9, value=d.get('calidad_detalle', '')).border = thin_border
            ws_sec.cell(row=r, column=10, value=d.get('fecha_publicacion', '')).border = thin_border
            ws_sec.cell(row=r, column=11, value=d.get('fecha_ultima_reforma', '')).border = thin_border
            ws_sec.cell(row=r, column=12, value='Sí' if d.get('es_escaneado') else 'No').border = thin_border
            ws_sec.cell(row=r, column=13, value=len(d.get('contenido', ''))).border = thin_border
            ws_sec.cell(row=r, column=14, value=d.get('url', '')).border = thin_border

            if (r - 2) % 2 == 1:
                for c in range(1, 15):
                    if ws_sec.cell(row=r, column=c).fill == PatternFill():
                        ws_sec.cell(row=r, column=c).fill = fill_alt
            r += 1

        ws_sec.column_dimensions['A'].width = 5
        ws_sec.column_dimensions['B'].width = 65
        ws_sec.column_dimensions['C'].width = 15
        ws_sec.column_dimensions['N'].width = 50
        ws_sec.freeze_panes = 'A2'
        ws_sec.auto_filter.ref = f"A1:N{r - 1}"

    ruta = os.path.join(BASE, "reporte_oaxaca.xlsx")
    wb.save(ruta)
    print(f"  Reporte guardado: {ruta}")


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":
    fix_jsons()
    generar_reporte()
