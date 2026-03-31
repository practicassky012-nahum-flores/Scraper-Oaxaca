#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Clasificador de documentos relevantes para litigio hipotecario.
Criterios estrictos:
  1. Códigos civiles (federal y estatal)
  2. Códigos de procedimientos civiles
  3. Leyes de vivienda
  4. Leyes de crédito / instituciones de crédito
  5. Leyes de títulos y operaciones de crédito
  6. Constituciones estatales
  7. Código de Comercio

Uso:
    python clasificar_hipotecario.py
"""

import os, json, shutil
from pathlib import Path
from datetime import datetime

BASE = Path(os.path.dirname(os.path.abspath(__file__)))
TEXTOS = BASE / "Textos de Oaxaca"
DEST = BASE / "hipotecario"
SECCIONES = ["Federal", "Estatal", "Municipal", "Marco Normativo"]

CAMPOS_V12 = [
    "titulo", "tipo_ordenamiento", "jurisdiccion", "materia",
    "fecha_publicacion", "es_vigente", "status_texto", "contenido",
    "url_fuente", "es_escaneado", "calidad_ocr", "archivo_origen"
]


def es_relevante(titulo, tipo):
    t = titulo.lower()
    razones = []
    if 'código civil' in t or ('civil' in t and tipo == 'Código'):
        razones.append('Código Civil')
    if 'procedimientos civiles' in t:
        razones.append('Procedimientos Civiles')
    if 'familiar' in t and ('código' in t or tipo == 'Código'):
        razones.append('Código Familiar')
    if 'vivienda' in t:
        razones.append('Vivienda')
    if any(x in t for x in ['instituciones de crédito', 'crédito', 'ahorro y crédito']):
        if 'información crediticia' not in t:
            razones.append('Crédito')
    if 'títulos y operaciones' in t:
        razones.append('Títulos y Operaciones de Crédito')
    if 'constitución' in t and 'estados unidos mexicanos' not in t:
        razones.append('Constitución Estatal')
    if 'comercio' in t and ('código' in t or tipo == 'Código'):
        razones.append('Código de Comercio')
    return razones


def tiene_12_campos(data):
    for campo in CAMPOS_V12:
        if campo not in data:
            return False
    return True


def calidad_general(data):
    """Bueno: tiene contenido largo + fecha + 12 campos.
    Regular: tiene contenido pero falta fecha o campos.
    Malo: sin contenido o muy corto."""
    contenido = data.get('contenido', '')
    tiene_campos = tiene_12_campos(data)
    tiene_fecha = data.get('fecha_publicacion') is not None
    tiene_texto = len(contenido) > 500

    if tiene_texto and tiene_fecha and tiene_campos:
        return 'Bueno'
    elif tiene_texto:
        return 'Regular'
    else:
        return 'Malo'


def main():
    print("=" * 70)
    print("  CLASIFICADOR HIPOTECARIO — OAXACA")
    print("=" * 70)

    DEST.mkdir(exist_ok=True)
    for f in DEST.iterdir():
        if f.suffix == '.json':
            f.unlink()

    total = 0
    relevantes = []

    for sec in SECCIONES:
        jd = TEXTOS / sec
        if not jd.exists():
            continue
        for jf in sorted(os.listdir(jd)):
            if not jf.endswith('.json'):
                continue
            total += 1
            with open(jd / jf, 'r', encoding='utf-8') as f:
                d = json.load(f)
            razones = es_relevante(d.get('titulo', ''), d.get('tipo_ordenamiento', ''))
            if razones:
                shutil.copy2(jd / jf, DEST / jf)
                relevantes.append({
                    'archivo': jf,
                    'seccion': sec,
                    'titulo': d.get('titulo', ''),
                    'tipo_ordenamiento': d.get('tipo_ordenamiento', ''),
                    'materia': d.get('materia', []),
                    'fecha': d.get('fecha_publicacion'),
                    'vigente': d.get('es_vigente'),
                    'razones': razones,
                    'tiene_12': tiene_12_campos(d),
                    'calidad': calidad_general(d),
                })

    with open(DEST / "indice_hipotecario.json", 'w', encoding='utf-8') as f:
        json.dump({
            "fecha": datetime.now().isoformat(),
            "total_analizados": total,
            "total_relevantes": len(relevantes),
            "documentos": relevantes
        }, f, ensure_ascii=False, indent=2)

    # Excel
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        alt = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        verde = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        rojo = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        amarillo = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        brd = Border(left=Side(style='thin', color='D9E2F3'), right=Side(style='thin', color='D9E2F3'),
                     top=Side(style='thin', color='D9E2F3'), bottom=Side(style='thin', color='D9E2F3'))

        ws = wb.active
        ws.title = "Hipotecario"
        ws.merge_cells('A1:K1')
        ws['A1'] = f"Litigio Hipotecario — Oaxaca ({len(relevantes)} de {total})"
        ws['A1'].font = Font(bold=True, size=14, color='2F5496')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = Font(size=10, italic=True, color='666666')

        headers = ['#', 'Estado', 'Sección', 'Nombre de JSON', 'Tipo de Ordenamiento',
                   'Materia', 'Fecha Publicación', 'Vigente', 'Razón de Relevancia',
                   '¿Tiene los 12 campos?', 'Calidad General']
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=c, value=h)
            cell.font = hf
            cell.fill = hfill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = brd

        for i, r in enumerate(relevantes, 1):
            row = i + 4
            ws.cell(row=row, column=1, value=i).border = brd
            ws.cell(row=row, column=2, value='Oaxaca').border = brd
            ws.cell(row=row, column=3, value=r['seccion']).border = brd
            ws.cell(row=row, column=4, value=r['titulo'][:100]).border = brd
            ws.cell(row=row, column=5, value=r['tipo_ordenamiento']).border = brd
            ws.cell(row=row, column=6, value=', '.join(r['materia'])).border = brd
            ws.cell(row=row, column=7, value=r['fecha'] or '').border = brd
            ws.cell(row=row, column=8, value='Sí' if r['vigente'] else 'No').border = brd
            ws.cell(row=row, column=9, value=', '.join(r['razones'])).border = brd

            # ¿Tiene los 12 campos?
            cell_12 = ws.cell(row=row, column=10, value='Sí' if r['tiene_12'] else 'No')
            cell_12.border = brd
            cell_12.alignment = Alignment(horizontal='center')
            cell_12.fill = verde if r['tiene_12'] else rojo

            # Calidad General
            cell_cal = ws.cell(row=row, column=11, value=r['calidad'])
            cell_cal.border = brd
            cell_cal.alignment = Alignment(horizontal='center')
            if r['calidad'] == 'Bueno':
                cell_cal.fill = verde
            elif r['calidad'] == 'Regular':
                cell_cal.fill = amarillo
            else:
                cell_cal.fill = rojo

            # Filas alternadas
            if i % 2 == 0:
                for c in range(1, 12):
                    cell = ws.cell(row=row, column=c)
                    if cell.fill == PatternFill():
                        cell.fill = alt

        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 55
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 18
        ws.column_dimensions['K'].width = 16
        ws.freeze_panes = 'A5'
        ws.auto_filter.ref = f"A4:K{len(relevantes) + 4}"

        wb.save(str(DEST / "reporte_hipotecario.xlsx"))
        print(f"  Reporte: {DEST / 'reporte_hipotecario.xlsx'}")
    except Exception as e:
        print(f"  Error Excel: {e}")

    print(f"\n  Analizados: {total} | Relevantes: {len(relevantes)}")
    for r in relevantes:
        campos_ok = "✓" if r['tiene_12'] else "✗"
        print(f"  {campos_ok} [{r['calidad']}] {r['titulo'][:50]}")
        print(f"    → {', '.join(r['razones'])}")
    print(f"\n  Copiados a: {DEST}")
    print("=" * 70)


if __name__ == "__main__":
    main()
