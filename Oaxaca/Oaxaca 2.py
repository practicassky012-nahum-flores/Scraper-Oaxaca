#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════╗
║     SCRAPER v8 — Leyes de Oaxaca & Leyes Federales                 ║
║     + OCR (pdfplumber / PaddleOCR / Tesseract) + JSON enriquecido ║
╚══════════════════════════════════════════════════════════════════════╝

Novedades v8:
  ✓ Título real extraído de la primera página del PDF (una sola pasada)
  ✓ Corrección de encoding para PDFs con fuentes no estándar
  ✓ JSON enriquecido: jurisdiccion, ordenamiento, fuente_oficial, status
  ✓ Nuevos campos: numero_paginas, es_escaneado, tiene_tablas
  ✓ Captura de fecha_pub y fecha_reforma desde tabla HTML
  ✓ Renombrado automático del JSON y PDF con el título real tras OCR
  ✓ OCR_WORKERS=1 para evitar crashes de pdfplumber en Windows

Dependencias:
  pip install pdfplumber pytesseract Pillow paddleocr paddlepaddle
  Tesseract binario (opcional): https://github.com/UB-Mannheim/tesseract/wiki
  PaddleOCR descarga modelos (~500 MB) la primera vez que se usa.
"""

import json
import re
import time
import warnings
import requests
from bs4 import BeautifulSoup
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import urljoin

# ── OCR imports (opcionales) ───────────────────────────────────────────────────
try:
    import pdfplumber
    _PDFPLUMBER_OK = True
except ImportError:
    _PDFPLUMBER_OK = False
    print("⚠ pdfplumber no instalado. Ejecuta: pip install pdfplumber")

try:
    import pytesseract
    _TESSERACT_OK = True
except ImportError:
    _TESSERACT_OK = False

try:
    from paddleocr import PaddleOCR
    import numpy as np
    _PADDLE_OK = True
except ImportError:
    _PADDLE_OK = False

_paddle_instance = None  # singleton lazy — el modelo (~500 MB) se carga una sola vez

warnings.filterwarnings("ignore", message="Unverified HTTPS request")

# ── Configuración ──────────────────────────────────────────────────────────────
BASE_DIR       = Path("Leyes de Oaxaca")
BASE_DIR_TEXTO = Path("Textos de Oaxaca")
MAX_WORKERS    = 10
OCR_WORKERS    = 1      # 1 hilo evita crashes de pdfplumber en Windows
DL_TIMEOUT     = 60

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

HEADERS_BASE = {
    "User-Agent":      UA,
    "Accept":          "text/html,application/xhtml+xml,*/*;q=0.9",
    "Accept-Language": "es-MX,es;q=0.9,en;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
}

SECCIONES = {
    "Federal": {
        "url":      "https://www.diputados.gob.mx/LeyesBiblio/index.htm",
        "pdf_base": "https://www.diputados.gob.mx/LeyesBiblio/",
        "tipo":     "federal",
    },
    "Estatal": {
        "url":            "https://www.congresooaxaca.gob.mx/legislaciones/legislacion_estatal.html",
        "tipo":           "oaxaca",
        "truncar_oaxaca": True,
    },
    "Municipal": {
        "url":            "https://www.congresooaxaca.gob.mx/legislaciones/legislacion_municipal.html",
        "tipo":           "oaxaca",
        "truncar_oaxaca": True,
    },
    "Marco Normativo": {
        "url":            "https://www.congresooaxaca.gob.mx/legislaciones/marco-normativo.html",
        "tipo":           "oaxaca",
        "truncar_oaxaca": False,
    },
}


# ── Utilidades de texto ────────────────────────────────────────────────────────

def sanitizar(texto: str, max_len: int = 160) -> str:
    """Nombre de archivo válido conservando acentos UTF-8."""
    texto = re.sub(r'\(Reformad[ao].*',            '', texto, flags=re.IGNORECASE | re.DOTALL)
    texto = re.sub(r'\(.*?[Ll]egislatura.*?\)',    '', texto, flags=re.DOTALL)
    texto = re.sub(r'\(.*?[Dd]ecreto.*?\)',        '', texto, flags=re.DOTALL)
    texto = re.sub(r'\(.*?[Aa]probad[ao].*?\)',    '', texto, flags=re.DOTALL)
    texto = re.sub(r'\(.*?[Rr]esoluci[oó]n.*?\)', '', texto, flags=re.DOTALL)
    for src, dst in [('\u2013', '-'), ('\u2014', '-'),
                     ('\u2018', ''),  ('\u2019', ''),
                     ('\u201c', ''),  ('\u201d', ''),
                     ('\u00b7', '.'), ('\u00ba', ''), ('\u00aa', '')]:
        texto = texto.replace(src, dst)
    texto = re.sub(r'[<>:"/\\|?*]', ' ', texto)
    texto = re.sub(r'[\x00-\x1f\x7f-\x9f]', '', texto)
    texto = re.sub(r"[^\w\s\-\.\,\(\)]", ' ', texto)
    texto = re.sub(r'\s+', ' ', texto).strip().rstrip('.,')
    return texto[:max_len] if texto else "sin_titulo"


def truncar_en_oaxaca(titulo: str) -> str:
    idx = titulo.lower().find("oaxaca")
    if idx != -1:
        return titulo[: idx + len("oaxaca")].strip().rstrip(".,")
    return titulo


def aplicar_nombre(num: int, titulo: str, seccion: str) -> str:
    """Nombre base del archivo: solo el título, sin número."""
    if seccion in ("Estatal", "Municipal"):
        titulo = truncar_en_oaxaca(titulo)
    return sanitizar(titulo, max_len=200)


# ── URLs de descarga ───────────────────────────────────────────────────────────

def preparar_urls(href: str, page_url: str, pdf_base: str = "") -> tuple[str, str]:
    if href.startswith("http"):
        primary = href
    elif pdf_base and not href.startswith(".."):
        primary = urljoin(pdf_base, href.lstrip("./"))
    else:
        primary = urljoin(page_url, href)
    m = re.match(r'^(https?://)([^/]+)/(docs\d+\.congresooaxaca\.gob\.mx/.+)$', primary)
    alt = (m.group(1) + m.group(3)) if m else ""
    return primary, alt


def es_pdf(href: str, img_src: str = "") -> bool:
    h = href.lower()
    i = img_src.lower()
    if h.endswith((".doc", ".docx", ".xls", ".xlsx")):
        return False
    if "_mov" in h or "pdf_mov" in h:
        return False
    return h.endswith(".pdf") or "pdf" in i


# ── Clasificación por materia ──────────────────────────────────────────────────

REGLAS_MATERIA: list[tuple[str, list[str]]] = [
    ("constitucional",          ["constitución", "constitucional"]),
    ("penal",                   ["código penal", "penal", "delito", "crimen",
                                  "readaptación social", "reinserción social",
                                  "ejecución de pena"]),
    ("civil",                   ["código civil", "familiar", "sucesiones",
                                  "arrendamiento", "propiedad privada", "notariado familiar"]),
    ("electoral",               ["electoral", "elecciones", "partido político",
                                  "sufragio", "electorales", "instituto electoral"]),
    ("fiscal",                  ["fiscal", "hacienda", "presupuesto", "impuesto",
                                  "tributario", "finanzas públicas", "ingresos del estado",
                                  "egresos", "deuda pública"]),
    ("laboral",                 ["laboral", "ley del trabajo", "empleo",
                                  "trabajador", "sindicato", "contrato colectivo",
                                  "servicio civil", "burocracia"]),
    ("derechos humanos",        ["derechos humanos", "derechos fundamentales",
                                  "comisión de derechos"]),
    ("transparencia",           ["transparencia", "acceso a la información",
                                  "datos personales", "protección de datos",
                                  "archivos", "gobierno abierto"]),
    ("género",                  ["género", "igualdad de género", "violencia contra la mujer",
                                  "mujeres", "feminicidio", "paridad",
                                  "hostigamiento sexual", "acoso sexual",
                                  "igualdad entre mujeres y hombres"]),
    ("niñez",                   ["niño", "niña", "niñez", "adolescente",
                                  "infancia", "menor de edad", "menores"]),
    ("salud",                   ["salud", "sanitario", "sanitaria", "médico",
                                  "médica", "enfermería", "hospital",
                                  "farmacéutico", "adicciones"]),
    ("educación",               ["educación", "educativa", "educativo",
                                  "universidad", "escolar", "escuela",
                                  "enseñanza", "bachillerato", "normal"]),
    ("cultura",                 ["cultura", "cultural", "patrimonio",
                                  "museo", "arte", "artístico", "lenguas indígenas",
                                  "patrimonio histórico"]),
    ("desarrollo social",       ["desarrollo social", "asistencia social",
                                  "bienestar social", "pobreza", "marginación",
                                  "asistencia pública"]),
    ("urbanismo",               ["desarrollo urbano", "ordenamiento territorial",
                                  "asentamiento", "vivienda", "construcción",
                                  "obra pública", "catastro"]),
    ("ambiental",               ["ambiente", "ambiental", "ecología", "ecológico",
                                  "recursos naturales", "forestal", "fauna",
                                  "flora", "biodiversidad", "sustentable",
                                  "residuos", "cambio climático"]),
    ("protección civil",        ["protección civil", "emergencia", "desastre",
                                  "riesgo", "gestión de riesgos", "bomberos"]),
    ("seguridad pública",       ["seguridad pública", "policía", "prevención del delito",
                                  "seguridad ciudadana", "fuerza pública",
                                  "guardia nacional"]),
    ("ciencia y tecnología",    ["ciencia", "tecnología", "innovación",
                                  "investigación científica", "tecnológico",
                                  "inteligencia artificial"]),
    ("deporte",                 ["deporte", "deportivo", "cultura física",
                                  "actividad física", "recreación", "olimpiada"]),
    ("municipal",               ["municipal", "municipio", "municipios",
                                  "cabildo", "ayuntamiento"]),
    ("turismo",                 ["turismo", "turística", "turístico"]),
    ("comunicación social",     ["radio", "televisión", "medios de comunicación",
                                  "comunicación social", "prensa", "periodismo"]),
    ("movilidad",               ["transporte", "vialidad", "tránsito",
                                  "tráfico", "movilidad", "carretera",
                                  "autopista", "aeropuerto"]),
    ("agrario",                 ["agrario", "agrícola", "rural", "campo",
                                  "campesino", "ganadero", "pesca",
                                  "silvícola", "ejido"]),
    ("económico",               ["económico", "economía", "competitividad",
                                  "industria", "comercio", "empresa",
                                  "inversión", "fomento económico"]),
    ("anticorrupción",          ["anticorrupción", "combate a la corrupción",
                                  "contraloría", "sistema anticorrupción"]),
    ("notariado",               ["notarial", "notario", "fedatario",
                                  "fe pública", "correduría"]),
    ("registral",               ["registro público", "registral",
                                  "registro civil", "registro de la propiedad"]),
    ("energía",                 ["energía", "energético", "eléctrico",
                                  "electricidad", "hidrocarburos", "gas"]),
    ("migratorio",              ["migración", "migrante", "migratorio",
                                  "refugiado", "desplazado", "apátrida"]),
    ("ética",                   ["ética", "código de conducta",
                                  "conducta de servidores", "conflicto de interés",
                                  "declaración patrimonial"]),
    ("parlamentario",           ["congreso", "legislativo", "parlamentario",
                                  "diputado", "senado", "cámara de diputados",
                                  "poder legislativo"]),
    ("auditoría",               ["auditoría", "fiscalización",
                                  "rendición de cuentas", "cuenta pública",
                                  "órgano de fiscalización", "auditoría superior"]),
    ("judicial",                ["poder judicial", "tribunal", "juzgado",
                                  "magistrado", "supremo tribunal", "juicio de amparo"]),
    ("participación ciudadana", ["participación ciudadana", "referéndum",
                                  "plebiscito", "consulta popular",
                                  "iniciativa popular"]),
    ("administrativo",          ["orgánica", "administrativa", "administrativo",
                                  "organización gubernamental",
                                  "servicio profesional de carrera",
                                  "servidor público", "función pública"]),
]


def clasificar_materia(titulo: str) -> list[str]:
    t = titulo.lower()
    encontradas = [m for m, kws in REGLAS_MATERIA if any(k in t for k in kws)]
    return encontradas if encontradas else ["general"]


def capitalizar_materia(materia: str) -> str:
    return " ".join(p.capitalize() for p in materia.split())


# ── Corrección de encoding ────────────────────────────────────────────────────

_CORRECCIONES_ENCODING: list[tuple[str, str]] = [
    ("CDIGO",         "Código"),
    ("CONSTITUCIN",   "Constitución"),
    ("LEGISLACIN",    "Legislación"),
    ("ADMINISTRACIN", "Administración"),
    ("REGLAMENTACIN", "Reglamentación"),
    ("REGULACIN",     "Regulación"),
    ("ORGANIZACIN",   "Organización"),
    ("COMUNICACIN",   "Comunicación"),
    ("EDUCACIN",      "Educación"),
    ("HABITACIN",     "Habitación"),
    ("INFORMACIN",    "Información"),
    ("PARTICIPACIN",  "Participación"),
    ("PROTECCIN",     "Protección"),
    ("SANCIN",        "Sanción"),
    ("FUNCIN",        "Función"),
    ("RELACIN",       "Relación"),
    ("GESTIN",        "Gestión"),
    ("CREACIN",       "Creación"),
    ("NACIN",         "Nación"),
    ("ACCIN",         "Acción"),
    ("APLICACIN",     "Aplicación"),
    ("PUBLICACIN",    "Publicación"),
    ("MODIFICACIN",   "Modificación"),
    ("DISPOSICIN",    "Disposición"),
    ("OBLIGACIN",     "Obligación"),
    ("DEFINICIN",     "Definición"),
    ("CONTAMINACIN",  "Contaminación"),
    ("PREVENCIN",     "Prevención"),
    ("ATENCIN",       "Atención"),
    ("CONSTRUCCIN",   "Construcción"),
    ("GENERACIN",     "Generación"),
    ("COOPERACIN",    "Cooperación"),
    ("RESOLUCIN",     "Resolución"),
    ("SELECCIN",      "Selección"),
    ("CONTRIBUCIN",   "Contribución"),
    ("EJECUCIN",      "Ejecución"),
    ("INSTITUCIN",    "Institución"),
    ("PBLICA",        "Pública"),
    ("PBLICO",        "Público"),
    ("MXICO",         "México"),
    ("REPBLICA",      "República"),
    ("DEMOCRTICA",    "Democrática"),
    ("JURDICA",       "Jurídica"),
    ("JURDICO",       "Jurídico"),
    ("ECONMICA",      "Económica"),
    ("ECONMICO",      "Económico"),
    ("ORGNICA",       "Orgánica"),
    ("ORGNICO",       "Orgánico"),
    ("POLTICA",       "Política"),
    ("POLTICO",       "Político"),
    ("TCNICA",        "Técnica"),
    ("TCNICO",        "Técnico"),
    ("HISTRICA",      "Histórica"),
    ("HISTRICO",      "Histórico"),
    ("TURSTICA",      "Turística"),
    ("TURSTICO",      "Turístico"),
    ("ENERGTICA",     "Energética"),
    ("ENERGTICO",     "Energético"),
    ("PERIDICO",      "Periódico"),
    ("MDICO",         "Médico"),
    ("MDICA",         "Médica"),
    ("AGRICOLA",      "Agrícola"),
    ("SILVICOLA",     "Silvícola"),
    ("CRITICA",       "Crítica"),
]

_RE_CORRECCIONES = [
    (re.compile(r'\b' + re.escape(m) + r'\b', re.IGNORECASE), b)
    for m, b in _CORRECCIONES_ENCODING
]


def corregir_encoding_comun(texto: str) -> str:
    """Corrige palabras legales mexicanas con encoding de fuente corrupto."""
    for patron, correcto in _RE_CORRECCIONES:
        texto = patron.sub(correcto, texto)
    return texto


# ── Title case inteligente para español ───────────────────────────────────────

_MINUSCULAS_ES = {
    "de", "del", "la", "los", "las", "el", "en", "con", "por", "para",
    "a", "ante", "bajo", "y", "o", "e", "u", "al", "que", "sobre",
    "sin", "sus", "su", "entre",
}


def _title_case_es(texto: str) -> str:
    palabras = texto.split()
    resultado = []
    for i, pal in enumerate(palabras):
        if i == 0 or pal.lower() not in _MINUSCULAS_ES:
            resultado.append(pal[0].upper() + pal[1:].lower() if pal else pal)
        else:
            resultado.append(pal.lower())
    return " ".join(resultado)


# ── Extracción unificada: título + texto en una sola pasada ──────────────────

_KEYWORDS_TITULO_LEGAL = [
    "LEY", "CÓDIGO", "CODIGO", "CONSTITUCIÓN", "CONSTITUCION",
    "REGLAMENTO", "DECRETO", "ACUERDO", "NORMA", "LINEAMIENTO",
    "CONVENCIÓN", "CONVENCION", "TRATADO", "CIRCULAR",
]


def _filtrar_lineas_titulo(lineas: list[str]) -> list[str]:
    """Devuelve el bloque de 1-5 líneas que forman el título legal."""
    resultado: list[str] = []
    en_titulo = False
    for linea in lineas:
        linea_up = linea.upper()
        es_kw = any(kw in linea_up for kw in _KEYWORDS_TITULO_LEGAL)
        if es_kw and not en_titulo:
            en_titulo = True
            resultado.append(linea)
        elif en_titulo:
            if (len(linea) > 5
                    and not re.match(r'^(Art[íi]culo|ARTÍCULO|Art\.)', linea, re.IGNORECASE)
                    and not re.match(r'^\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', linea)
                    and not re.match(r'^(Página|PÁGINA|\d+\s*$)', linea, re.IGNORECASE)):
                resultado.append(linea)
                if len(resultado) >= 5:
                    break
            else:
                break
    return resultado


def _titulo_desde_texto(texto_p0: str) -> str:
    """Extrae título de texto pdfplumber aplicando correcciones de encoding."""
    texto_corr = corregir_encoding_comun(texto_p0)
    lineas = [l.strip() for l in texto_corr.splitlines() if l.strip()][:25]
    tl = _filtrar_lineas_titulo(lineas)
    if tl:
        return _title_case_es(re.sub(r'\s+', ' ', " ".join(tl)).strip())
    primeras = [l for l in lineas if len(l) > 10][:3]
    if primeras:
        return _title_case_es(re.sub(r'\s+', ' ', " ".join(primeras)).strip())
    return ""


def _get_paddle() -> "PaddleOCR":
    """Singleton lazy de PaddleOCR (carga el modelo una sola vez)."""
    global _paddle_instance
    if _paddle_instance is None:
        _paddle_instance = PaddleOCR(
            use_angle_cls=True,   # corrige páginas rotadas
            lang="es",            # español
            use_gpu=False,
            show_log=False,
        )
    return _paddle_instance


def _paddle_ocr_pagina(img) -> str:
    """
    Recibe una imagen PIL y devuelve el texto extraído por PaddleOCR,
    ordenado de arriba-abajo / izquierda-derecha.
    Solo incluye fragmentos con confianza > 0.5.
    """
    ocr = _get_paddle()
    img_arr = np.array(img)
    resultado = ocr.ocr(img_arr, cls=True)
    if not resultado or not resultado[0]:
        return ""
    fragmentos = []
    for linea in resultado[0]:
        bbox, (texto, confianza) = linea
        if confianza > 0.5 and texto.strip():
            y_top = bbox[0][1]   # coordenada vertical del borde superior
            x_left = bbox[0][0]
            fragmentos.append((y_top, x_left, texto.strip()))
    fragmentos.sort(key=lambda t: (t[0], t[1]))
    return "\n".join(t[2] for t in fragmentos)


def _procesar_pdf_completo(pdf_path: Path) -> dict:
    """
    Abre pdfplumber UNA SOLA VEZ y extrae en la misma pasada:
      titulo_real, texto, numero_paginas, es_escaneado, tiene_tablas, metodo
    """
    _VACIO = {
        "titulo_real": "", "texto": "", "numero_paginas": 0,
        "es_escaneado": False, "tiene_tablas": False,
        "metodo": "pdfplumber no instalado",
    }
    if not _PDFPLUMBER_OK:
        return _VACIO

    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            num_paginas = len(pdf.pages)
            if num_paginas == 0:
                return {**_VACIO, "metodo": "PDF vacío"}

            partes: list[str] = []
            pags_ocr = 0
            tiene_tablas = False
            titulo_real = ""

            for i, page in enumerate(pdf.pages):
                texto_pag = (page.extract_text() or "").strip()

                # Heurística rápida de tablas (primeras 5 páginas)
                if i < 5 and not tiene_tablas:
                    lineas_pag = texto_pag.splitlines()
                    if sum(1 for l in lineas_pag if l.count("  ") >= 2 or "\t" in l) >= 3:
                        tiene_tablas = True

                # Página 0: extraer título
                # FIX: la imagen se genera UNA sola vez y se reutiliza para texto OCR
                if i == 0:
                    img = None
                    try:
                        if _PADDLE_OK or _TESSERACT_OK:
                            img = page.to_image(resolution=200).original
                    except Exception as exc_img:
                        warnings.warn(f"No se pudo convertir página 0 a imagen: {exc_img}")

                    if _PADDLE_OK and img is not None:
                        try:
                            raw_paddle = _paddle_ocr_pagina(img)
                            lineas_ocr = [l.strip() for l in raw_paddle.splitlines() if l.strip()][:25]
                            tl = _filtrar_lineas_titulo(lineas_ocr)
                            if tl:
                                titulo_real = _title_case_es(
                                    re.sub(r'\s+', ' ', " ".join(tl)).strip()
                                )
                        except (RuntimeError, ValueError, AttributeError) as exc_paddle:
                            warnings.warn(f"PaddleOCR falló en título: {exc_paddle}")
                    elif _TESSERACT_OK and img is not None:
                        try:
                            raw_ocr = pytesseract.image_to_string(
                                img, lang="spa", config="--psm 6"
                            )
                            lineas_ocr = [l.strip() for l in raw_ocr.splitlines() if l.strip()][:25]
                            tl = _filtrar_lineas_titulo(lineas_ocr)
                            if tl:
                                titulo_real = _title_case_es(
                                    re.sub(r'\s+', ' ', " ".join(tl)).strip()
                                )
                        except (RuntimeError, OSError, AttributeError) as exc_tess:
                            warnings.warn(f"Tesseract falló en título: {exc_tess}")

                    if not titulo_real and texto_pag:
                        titulo_real = _titulo_desde_texto(texto_pag)

                # Texto de la página
                # FIX: img ya existe para página 0; solo se genera de nuevo en páginas siguientes
                if len(texto_pag) >= 30:
                    partes.append(texto_pag)
                elif _PADDLE_OK:
                    try:
                        if i != 0:
                            img = page.to_image(resolution=200).original
                        if img is not None:
                            texto_ocr = _paddle_ocr_pagina(img)
                            if texto_ocr:
                                partes.append(f"[Pág {i+1} — OCR]\n{texto_ocr}")
                                pags_ocr += 1
                    except (RuntimeError, ValueError, AttributeError) as exc_paddle:
                        warnings.warn(f"PaddleOCR falló en pág {i+1}: {exc_paddle}")
                elif _TESSERACT_OK:
                    try:
                        if i != 0:
                            img = page.to_image(resolution=200).original
                        if img is not None:
                            texto_ocr = pytesseract.image_to_string(
                                img, lang="spa", config="--psm 3"
                            ).strip()
                            if texto_ocr:
                                partes.append(f"[Pág {i+1} — OCR]\n{texto_ocr}")
                                pags_ocr += 1
                    except (RuntimeError, OSError, AttributeError) as exc_tess:
                        warnings.warn(f"Tesseract falló en pág {i+1}: {exc_tess}")

            contenido = "\n\n".join(partes)
            es_escaneado = pags_ocr / num_paginas > 0.5
            # FIX: motor refleja el que realmente se usó, no el que está instalado
            if pags_ocr == 0:
                metodo = "pdfplumber"
            elif _PADDLE_OK:
                metodo = f"pdfplumber+paddle ({pags_ocr} págs OCR)"
            else:
                metodo = f"pdfplumber+tesseract ({pags_ocr} págs OCR)"

            return {
                "titulo_real":    titulo_real,
                "texto":          contenido,
                "numero_paginas": num_paginas,
                "es_escaneado":   es_escaneado,
                "tiene_tablas":   tiene_tablas,
                "metodo":         metodo,
            }

    except Exception as e:
        return {**_VACIO, "metodo": f"error: {str(e)[:80]}"}


# ── Metadatos legales ─────────────────────────────────────────────────────────

_ORDENAMIENTOS: list[tuple[str, str]] = [
    ("constitución", "Constitución"),
    ("código",       "Código"),
    ("reglamento",   "Reglamento"),
    ("decreto",      "Decreto"),
    ("acuerdo",      "Acuerdo"),
    ("convención",   "Convención"),
    ("tratado",      "Tratado"),
    ("circular",     "Circular"),
    ("lineamiento",  "Lineamiento"),
    ("norma",        "Norma"),
    ("ley",          "Ley"),
]


def detectar_ordenamiento(titulo: str, tipo_tabla: str = "") -> str:
    t = titulo.lower()
    for kw, label in _ORDENAMIENTOS:
        if kw in t:
            return label
    return tipo_tabla if tipo_tabla else "Ley"


def obtener_jurisdiccion(nombre_seccion: str) -> str:
    return f"Congreso del Estado de Oaxaca - {nombre_seccion}"


def obtener_fuente_oficial(nombre_seccion: str) -> str:
    if nombre_seccion == "Federal":
        return "Cámara de Diputados del H. Congreso de la Unión"
    return "Congreso del Estado de Oaxaca"


def detectar_status(contenido: str) -> str:
    muestra = contenido[:2000].upper()
    if "ABROGADO" in muestra:
        return "Abrogado"
    if "DEROGADO" in muestra:
        return "Derogado"
    return "Vigente"


def _clasificar_materia_nueva(titulo: str) -> list:
    """Clasifica materia con las 38 categorías del proyecto. Basado solo en el título."""
    t = titulo.lower()
    materias = []

    # 1. Constitucional
    if 'constitución' in t or 'constitucional' in t:
        materias.append('constitucional')
    # 2. Civil
    if any(x in t for x in ['código civil', 'civil', 'familiar', 'sucesión', 'sucesiones', 'matrimonio', 'divorcio', 'adopción', 'adopciones', 'propiedad', 'condominio']):
        if 'protección civil' not in t and 'servicio civil' not in t and 'registro civil' not in t:
            materias.append('civil')
    # 3. Penal
    if any(x in t for x in ['penal', 'delito', 'delincuencia', 'ejecución de penas', 'ejecución penal', 'ejecución de sanciones', 'reinserción', 'prisión', 'control de sustancias', 'control de precursores', 'prevención e identificación de operaciones', 'tortura', 'juegos y sorteos']):
        materias.append('penal')
    # 4. Administrativo
    if any(x in t for x in ['administrativ', 'administración pública', 'adquisiciones', 'obra pública', 'obras públicas', 'mejora regulatoria', 'archivos', 'austeridad', 'entrega-recepción', 'entrega recepción', 'servicio civil', 'bienes', 'remuneración', 'servidor público', 'servidores públicos', 'entidades paraestatales', 'expropiación', 'firma electrónica', 'trámite', 'burocrát', 'orgánica de', 'orgánica del', 'nacionalidad', 'estatuto de gobierno', 'ejercicio profesional', 'consejo consultivo', 'patronato', 'fideicomiso', 'extradición', 'cooperación internacional', 'servicio exterior', 'convenio constitutivo', 'adhesión de méxico', 'religiosa', 'culto', 'estadístic', 'geográfic', 'husos horarios', 'escudo', 'bandera', 'himno', 'planeación', 'desarrollo administrativo']):
        materias.append('administrativo')
    # 5. Fiscal
    if any(x in t for x in ['fiscal', 'impuesto', 'hacendaria', 'hacienda', 'ingreso', 'egreso', 'presupuesto', 'deuda pública', 'coordinación fiscal', 'catastro', 'contribución', 'contribuyente', 'crédito', 'bancari', 'financier', 'ahorro', 'seguro social', 'fianza', 'valores', 'mercado de valores', 'moneda', 'monetari', 'hipotecar', 'tesorería', 'tributar', 'servicio de administración tributaria', 'unidad de medida', 'contabilidad gubernamental', 'derechos del estado', 'federal de derechos', 'sistemas de pago', 'banco de méxico', 'protección al ahorro']):
        materias.append('fiscal')
    # 6. Laboral
    if any(x in t for x in ['trabajo', 'trabajador', 'laboral', 'salario', 'sindicat', 'empleo', 'servicio profesional', 'pensiones para los integrantes']):
        materias.append('laboral')
    # 7. Electoral
    if any(x in t for x in ['electoral', 'elección', 'elecciones', 'partido político', 'partidos', 'voto', 'revocación de mandato', 'proceso electoral']):
        materias.append('electoral')
    # 8. Derechos Humanos
    if any(x in t for x in ['derechos humanos', 'víctima', 'víctimas', 'discriminación', 'no discriminación', 'igualdad', 'desaparición', 'desplazamiento', 'amnistía', 'personas adultas mayores', 'defensoría', 'persona desaparecida', 'declaración especial de ausencia', 'localización, recuperación', 'violencia y acoso', 'indígena', 'pueblo', 'comunidad indígena', 'discapacidad', 'inclusión']):
        if 'igualdad de género' not in t and 'igualdad entre' not in t:
            materias.append('derechos humanos')
    # 9. Transparencia
    if any(x in t for x in ['transparencia', 'acceso a la información', 'datos personales', 'información pública', 'protección de datos', 'clasificación y desclasificación', 'lineamientos técnicos generales para la publicación']):
        materias.append('transparencia')
    # 10. Género
    if any(x in t for x in ['género', 'mujeres', 'violencia contra la mujer', 'violencia de género', 'igualdad de género', 'igualdad entre mujeres', 'vida libre de violencia']):
        materias.append('género')
    # 11. Derechos de la Niñez
    if any(x in t for x in ['niñas', 'niños', 'adolescentes', 'niñez', 'menores', 'menor de edad', 'paternidad responsable']):
        materias.append('derechos de la niñez')
    # 12. Salud
    if any(x in t for x in ['salud', 'sanitari', 'médic', 'hospital', 'fumador', 'tabaco', 'drogas', 'adiccion', 'mental', 'cuidados paliativos', 'vih', 'sida', 'voluntad anticipada', 'donación', 'cáncer', 'espectro autista']):
        materias.append('salud')
    # 13. Educación
    if any(x in t for x in ['educación', 'educativ', 'escuela', 'universidad', 'enseñanza', 'maestro', 'alumno', 'becas', 'lectura', 'bachiller', 'colegio de bachilleres', 'instituto de estudios', 'instituto politécnico', 'academia', 'lengua mixteca', 'lengua zapoteca', 'instituto de investigacion']):
        materias.append('educación')
    # 14. Cultura
    if any(x in t for x in ['cultura', 'cultural', 'patrimonio cultural', 'arte', 'museo', 'biblioteca', 'artístic', 'fomento a la lectura', 'cinematografí', 'derecho de autor', 'antropología', 'festividad', 'software libre', 'premios']):
        materias.append('cultura')
    # 15. Desarrollo Social
    if any(x in t for x in ['desarrollo social', 'asistencia social', 'pobreza', 'bienestar', 'beneficencia', 'alimenta', 'vivienda', 'habitacional', 'población', 'prestación de servicios para la atención', 'desarrollo integral de la familia', 'promotores voluntarios', 'integración social', 'personas jóvenes', 'juventud', 'sociedades de solidaridad']):
        materias.append('desarrollo social')
    # 16. Urbanismo
    if any(x in t for x in ['urbano', 'territorial', 'asentamientos humanos', 'ordenamiento territorial', 'desarrollo urbano', 'fraccionamiento', 'construcción', 'edificación', 'pavimentación', 'planificación y urbanización']):
        materias.append('urbanismo')
    # 17. Ambiental
    if any(x in t for x in ['ambiente', 'ambiental', 'ecológ', 'ecología', 'forestal', 'cambio climático', 'residuo', 'sustentable', 'biodiversidad', 'fauna', 'flora', 'animal', 'contaminación', 'vida silvestre', 'sanidad vegetal', 'variedades vegetales', 'vertimiento', 'zonas marinas']):
        materias.append('ambiental')
    # 18. Protección Civil
    if any(x in t for x in ['protección civil', 'emergencia', 'desastre', 'bombero', 'riesgo']):
        materias.append('protección civil')
    # 19. Seguridad Pública
    if any(x in t for x in ['seguridad pública', 'policía', 'guardia', 'prevención del delito', 'uso de la fuerza', 'seguridad nacional', 'seguridad interior', 'seguridad privada', 'extinción de dominio', 'detención', 'detenciones', 'armas de fuego', 'explosivos', 'armada', 'ejército', 'fuerza aérea', 'militar', 'servicio militar', 'disciplina para el personal', 'ascensos de la armada', 'neutralidad']):
        materias.append('seguridad pública')
    # 20. Ciencia y Tecnología
    if any(x in t for x in ['ciencia', 'tecnología', 'innovación', 'investigación científica', 'agencia espacial']):
        materias.append('ciencia y tecnología')
    # 21. Deporte
    if any(x in t for x in ['deporte', 'cultura física', 'deportiv']):
        materias.append('deporte')
    # 22. Municipal
    if any(x in t for x in ['municipal', 'municipio', 'ayuntamiento', 'cabildo']):
        materias.append('municipal')
    # 23. Turismo
    if any(x in t for x in ['turismo', 'turístic', 'turista']):
        materias.append('turismo')
    # 24. Comunicación Social
    if any(x in t for x in ['radio', 'televisión', 'telecomunicacion', 'radiodifusión', 'medios de comunicación', 'comunicación social', 'vías generales de comunicación', 'servicio postal']):
        materias.append('comunicación social')
    # 25. Movilidad
    if any(x in t for x in ['transporte', 'tránsito', 'vialidad', 'movilidad', 'camino', 'carretera', 'ferroviar', 'vehículo', 'aeropuerto', 'puerto', 'espacio aéreo', 'mar ']):
        materias.append('movilidad')
    # 26. Agrario
    if any(x in t for x in ['agrari', 'rural', 'campo', 'ejid', 'agríco', 'ganad', 'pecuari', 'maíz nativo', 'productos orgánicos', 'biocombustible']):
        materias.append('agrario')
    # 27. Económico
    if any(x in t for x in ['económic', 'competencia', 'comercio', 'inversión', 'industria', 'fomento económico', 'desarrollo económico', 'asociaciones público privadas', 'competitividad', 'economía circular', 'emprendimiento', 'mercantil', 'cooperativ', 'concurso', 'cámara empresarial', 'sociedades', 'consumidor', 'infraestructura', 'calidad', 'denominación', 'fomento', 'economía social', 'empresa pública', 'monte de piedad', 'casas de empeño', 'minería', 'aduaner']):
        materias.append('económico')
    # 28. Anticorrupción
    if any(x in t for x in ['anticorrupción', 'corrupción', 'contraloría', 'responsabilidades administrativas', 'responsabilidad patrimonial']):
        materias.append('anticorrupción')
    # 29. Notariado
    if any(x in t for x in ['notari', 'fe pública', 'correduría']):
        materias.append('notariado')
    # 30. Registral
    if any(x in t for x in ['registro público', 'registro civil', 'registral', 'catastro']):
        materias.append('registral')
    # 31. Energía
    if any(x in t for x in ['energía', 'eléctric', 'petróleo', 'hidrocarburos', 'geotermia', 'comisión federal de electricidad']):
        materias.append('energía')
    # 32. Migratorio
    if any(x in t for x in ['migra', 'refugiado', 'asilo', 'extranjero', 'movilidad humana']):
        materias.append('migratorio')
    # 33. Ética
    if any(x in t for x in ['ética', 'código de conducta', 'conducta']):
        materias.append('ética')
    # 34. Parlamentario
    if any(x in t for x in ['congreso', 'legislat', 'parlamentar', 'cámara de diputados', 'senado', 'poder legislativo', 'orgánica del poder legislativo']):
        materias.append('parlamentario')
    # 35. Auditoría
    if any(x in t for x in ['auditoría', 'fiscalización', 'cuenta pública', 'rendición de cuentas']):
        materias.append('auditoría')
    # 36. Judicial
    if any(x in t for x in ['tribunal', 'judicial', 'judicatura', 'justicia', 'poder judicial', 'juicio político', 'magistrat', 'mediación', 'controversia', 'conciliación']):
        materias.append('judicial')
    # 37. Participación Ciudadana
    if any(x in t for x in ['participación ciudadana', 'consulta popular', 'plebiscito', 'referéndum']):
        materias.append('participación ciudadana')
    # 38. General (fallback)
    return materias if materias else ['general']


def _detectar_tipo_ordenamiento(titulo: str) -> str:
    """Detecta tipo de ordenamiento. Solo 8 valores permitidos."""
    t = titulo.lower()
    if 'constitución' in t or 'constitucion' in t: return 'Constitución'
    if 'código' in t or 'codigo' in t: return 'Código'
    if 'reglamento' in t: return 'Reglamento'
    if 'decreto' in t: return 'Decreto'
    if 'acuerdo' in t: return 'Acuerdo'
    if 'nom-' in t or 'norma oficial' in t: return 'NOM'
    if 'ley' in t: return 'Ley'
    return 'Otro'


def _fecha_a_iso(fecha_str: str):
    """Convierte DD/MM/YYYY o DD-MM-YYYY → YYYY-MM-DD. Retorna None si no se puede."""
    if not fecha_str:
        return None
    # DD/MM/YYYY o DD-MM-YYYY
    m = re.match(r'(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})', fecha_str.strip())
    if m:
        return f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"
    # Ya viene en ISO YYYY-MM-DD
    m2 = re.match(r'(\d{4})-(\d{2})-(\d{2})', fecha_str.strip())
    if m2:
        return fecha_str.strip()
    return None


def _determinar_vigencia(titulo: str, contenido_inicio: str):
    """Solo mira título y encabezado. NO infiere de menciones internas."""
    t = titulo.upper()
    c = contenido_inicio[:500].upper() if contenido_inicio else ""
    if re.search(r'\bABROGAD[OA]\b', t): return False, "Abrogada"
    if re.search(r'\bDEROGAD[OA]\b', t): return False, "Derogada"
    if re.search(r'\bABROGAD[OA]\b', c): return False, "Abrogada"
    if re.search(r'\bDEROGAD[OA]\b', c): return False, "Derogada"
    return True, "Vigente"


def _calcular_calidad_v8(contenido: str, numero_paginas: int, metodo: str) -> dict:
    """Calcula calidad del texto extraído (0-100) — fórmula v8 estándar."""
    if not contenido or not contenido.strip():
        return {"calidad_porcentaje": 0, "calidad_detalle": "Sin contenido"}
    chars = len(contenido)
    paginas = max(numero_paginas, 1)
    score_densidad = min((chars / paginas) / 2000, 1.0)
    alfanum = sum(1 for c in contenido if c.isalnum() or c in ' \n\t.,;:()áéíóúñüÁÉÍÓÚÑÜ"\'—–-')
    ratio_alfanum = alfanum / max(chars, 1)
    lineas = contenido.split('\n')
    lineas_con_texto = sum(1 for l in lineas if len(l.strip()) > 5)
    ratio_lineas = lineas_con_texto / max(len(lineas), 1)
    palabras = contenido.split()
    palabras_cortas = sum(1 for p in palabras if len(p) == 1 and p.lower() not in 'aeiouy')
    ratio_basura = 1.0 - min(palabras_cortas / max(len(palabras), 1) * 5, 0.5)
    calidad = score_densidad * 0.30 + ratio_alfanum * 0.25 + ratio_lineas * 0.20 + ratio_basura * 0.25
    porcentaje = round(min(calidad * 100, 100))
    if porcentaje >= 85:   detalle = "Excelente"
    elif porcentaje >= 70: detalle = "Buena"
    elif porcentaje >= 50: detalle = "Aceptable"
    elif porcentaje >= 30: detalle = "Baja"
    else:                  detalle = "Muy baja"
    return {"calidad_porcentaje": porcentaje, "calidad_detalle": detalle}


def _limpiar_titulo(titulo: str) -> str:
    """Limpia el título: quita '(Reformada mediante...)', DOF, legislatura, etc."""
    # Quitar todo desde "(Reformad..." en adelante
    titulo = re.sub(r'\s*\(Reformad[ao].*', '', titulo, flags=re.IGNORECASE | re.DOTALL)
    # Quitar "(Decreto..."
    titulo = re.sub(r'\s*\(Decreto.*', '', titulo, flags=re.IGNORECASE | re.DOTALL)
    # Quitar "(Aprobad..."
    titulo = re.sub(r'\s*\(Aprobad[ao].*', '', titulo, flags=re.IGNORECASE | re.DOTALL)
    # Quitar "(Abrogad..."
    titulo = re.sub(r'\s*\(Abrogad[ao].*', '', titulo, flags=re.IGNORECASE | re.DOTALL)
    # Quitar "(vigente a partir..."
    titulo = re.sub(r'\s*\(vigente\s+a\s+partir.*', '', titulo, flags=re.IGNORECASE | re.DOTALL)
    # Quitar DOF DD/MM/YYYY o DD-MM-YYYY
    titulo = re.sub(r'\s*DOF\s+\d{2}[/\-]\d{2}[/\-]\d{4}.*', '', titulo, flags=re.IGNORECASE)
    # Quitar punto final
    titulo = titulo.rstrip('. ')
    # Limpiar espacios
    titulo = re.sub(r'\s+', ' ', titulo).strip()
    return titulo


# ── Guardar JSON enriquecido ──────────────────────────────────────────────────

def _extraer_fecha_publicacion_del_contenido(texto: str):
    """Busca TODAS las fechas de publicación en el Periódico Oficial y retorna la MÁS ANTIGUA.

    Formatos reconocidos:
      - "publicado en el Periódico Oficial... de fecha DD de MES de YYYY"
      - "Código publicado en la Segunda Sección del Periódico Oficial... el DD de MES de YYYY"
      - "Periódico Oficial... DD de MES de YYYY"
      - "P.O. del DD de MES de YYYY"
    """
    if not texto:
        return None
    meses = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
    }
    texto_lower = texto[:30000].lower().replace('\n', ' ')

    # Buscar TODAS las fechas asociadas a "Periódico Oficial" o "P.O."
    patrones = [
        # "publicado/a en ... Periódico Oficial ... DD de MES de YYYY"
        r'publicad[oa]?\s+en\s+[^\n]{0,250}?peri[oó]dico\s+oficial[^\n]{0,250}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:del?\s+)?(\d{4})',
        # "publicado/a en el Diario Oficial de la Federación el DD de MES de YYYY"
        r'publicad[oa]?\s+en\s+[^\n]{0,250}?diario\s+oficial\s+de\s+la\s+federaci[oó]n[^\n]{0,250}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:del?\s+)?(\d{4})',
        # "Diario Oficial de la Federación ... DD de MES de YYYY"
        r'diario\s+oficial\s+de\s+la\s+federaci[oó]n[^\n]{0,250}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:del?\s+)?(\d{4})',
        # "Periódico Oficial ... DD de MES de YYYY"
        r'peri[oó]dico\s+oficial[^\n]{0,250}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:del?\s+)?(\d{4})',
        # "P.O. ... DD de MES de YYYY" o "D.O.F. ... DD de MES de YYYY"
        r'(?:p\.?\s*o\.?|d\.?\s*o\.?\s*f\.?)\s+[^\n]{0,150}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+(?:del?\s+)?(\d{4})',
    ]

    todas_fechas = []
    for patron in patrones:
        for m in re.finditer(patron, texto_lower):
            try:
                dia = int(m.group(1))
                mes_str = m.group(2)
                anio = int(m.group(3))
                if 1800 < anio < 2100 and 1 <= dia <= 31 and mes_str in meses:
                    fecha_iso = f"{anio}-{meses[mes_str]}-{str(dia).zfill(2)}"
                    todas_fechas.append((anio, int(meses[mes_str]), dia, fecha_iso))
            except (ValueError, IndexError):
                continue

    if not todas_fechas:
        return None

    # Retornar la MÁS ANTIGUA
    todas_fechas.sort()
    return todas_fechas[0][3]


def guardar_json_texto(doc: dict, nombre_seccion: str,
                       extraccion: dict, titulo_real: str,
                       json_path: Path) -> None:
    # Título = nombre del archivo (siempre coinciden)
    titulo_final = json_path.stem

    texto = extraccion["texto"]
    es_escaneado = extraccion["es_escaneado"]

    # Calidad OCR (solo si es escaneado)
    calidad = _calcular_calidad_v8(texto, extraccion["numero_paginas"], extraccion.get("metodo", ""))
    calidad_ocr = calidad["calidad_porcentaje"] if es_escaneado else None

    # Vigencia
    es_vigente, status_texto = _determinar_vigencia(titulo_final, texto)

    # Fecha: extraer del contenido del PDF (buscar "publicado en el Periódico Oficial...")
    fecha_pub = _extraer_fecha_publicacion_del_contenido(texto)

    # archivo_origen = nombre del PDF
    archivo_origen = json_path.stem + ".pdf"

    datos = {
        "titulo":              titulo_final,
        "tipo_ordenamiento":   _detectar_tipo_ordenamiento(titulo_final),
        "jurisdiccion":        "Federal" if nombre_seccion == "Federal" else "Oaxaca",
        "materia":             _clasificar_materia_nueva(titulo_final),
        "fecha_publicacion":   fecha_pub,
        "es_vigente":          es_vigente,
        "status_texto":        status_texto,
        "contenido":           texto,
        "url_fuente":          doc.get("url", None) or None,
        "es_escaneado":        es_escaneado,
        "calidad_ocr":         calidad_ocr,
        "archivo_origen":      archivo_origen,
    }

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(datos, f, ensure_ascii=False, indent=2)


# ── Extracción + guardado + renombrado ────────────────────────────────────────

def _extraer_y_guardar(pdf_path: Path, json_path: Path,
                       doc: dict, nombre_seccion: str) -> tuple[bool, str]:
    """Extrae título + texto, guarda JSON y lo renombra inmediatamente con el título real."""
    resultado   = _procesar_pdf_completo(pdf_path)
    titulo_real = resultado.pop("titulo_real", "")
    guardar_json_texto(doc, nombre_seccion, resultado, titulo_real, json_path)

    # Renombrar JSON y PDF con el título limpio
    titulo_final = _limpiar_titulo(doc["titulo"])
    stem_nuevo = aplicar_nombre(doc['num'], titulo_final, nombre_seccion)
    nuevo_json   = json_path.parent / f"{stem_nuevo}.json"
    if nuevo_json != json_path and not nuevo_json.exists():
        try:
            json_path.rename(nuevo_json)
            nuevo_pdf = pdf_path.parent / f"{stem_nuevo}.pdf"
            if not nuevo_pdf.exists():
                pdf_path.rename(nuevo_pdf)
        except Exception:
            pass

    return bool(resultado["texto"].strip()), resultado["metodo"]


def _any_file(carpeta: Path, num: int, ext: str, nombre_base: str = "") -> "Path | None":
    """Busca archivo existente por número '{num:03d} - *.ext' o por nombre exacto."""
    # Primero buscar por nombre exacto (Federal usa título sin número)
    if nombre_base:
        exacto = carpeta / f"{nombre_base}.{ext}"
        if exacto.exists():
            return exacto
    # Luego buscar por número (secciones Oaxaca)
    hits = sorted(carpeta.glob(f"{num:03d} - *.{ext}"))
    return hits[0] if hits else None


def procesar_textos_seccion(nombre: str, docs: list, carpeta_pdfs: Path) -> None:
    """Para cada PDF descargado, extrae texto, guarda JSON y renombra."""
    if not _PDFPLUMBER_OK:
        print("  ⚠ pdfplumber no disponible — omitiendo extracción de texto.")
        return

    json_dir = BASE_DIR_TEXTO / nombre
    json_dir.mkdir(parents=True, exist_ok=True)

    tareas: list[tuple] = []
    omitidos = 0

    for d in docs:
        base      = aplicar_nombre(d["num"], d["titulo"], nombre)
        json_path = json_dir / f"{base}.json"

        # Skip si ya existe JSON con ese nombre o número
        if _any_file(json_dir, d["num"], "json", nombre_base=base):
            omitidos += 1
            continue

        # Usar el PDF real (por nombre exacto o por número)
        pdf_path = _any_file(carpeta_pdfs, d["num"], "pdf", nombre_base=base)
        if pdf_path is None or pdf_path.stat().st_size < 1024:
            continue

        tareas.append((pdf_path, json_path, d, nombre))

    if omitidos:
        print(f"  ↩  {omitidos} JSON ya existen — omitidos")
    if not tareas:
        print("  ✓  Sin textos nuevos que extraer.")
        return

    print(f"  → Extrayendo texto de {len(tareas)} PDFs ({OCR_WORKERS} hilo)\n")

    ok = fail = 0
    with ThreadPoolExecutor(max_workers=OCR_WORKERS) as ex:
        futuros = {
            ex.submit(_extraer_y_guardar, pdf_p, json_p, doc, nom): pdf_p.name
            for pdf_p, json_p, doc, nom in tareas
        }
        for fut in as_completed(futuros):
            nombre_pdf = futuros[fut]
            try:
                exito, metodo = fut.result()
                if exito:
                    ok += 1
                    print(f"  ✓  {nombre_pdf}  [{metodo}]", flush=True)
                else:
                    fail += 1
                    print(f"  ⚠  {nombre_pdf}  [sin texto — {metodo}]", flush=True)
            except Exception as e:
                fail += 1
                print(f"  ✗  {nombre_pdf}  [{e}]", flush=True)

    print(f"\n  OCR [{nombre}]: {ok} con texto, {fail} vacíos/error\n")


# ── Detección de columnas en tabla HTML ───────────────────────────────────────

KEYWORDS_OK = ["DENOMINACION", "TITULO DE LA LEY", "TITULO", "LEY", "NORMA", "NOMBRE"]
KEYWORDS_NO = ["TIPO", "FECHA", "DESCARGA", "PUBLICACION", "REFORMA", "NO."]
DATE_RE     = re.compile(r'^\d{2,4}[-/]\d{2}[-/]\d{2,4}$')


def detectar_col_titulo(headers: list, rows_sample: list) -> int:
    for kw in KEYWORDS_OK:
        for i, h in enumerate(headers):
            h_up = h.upper()
            if kw in h_up and not any(k in h_up for k in KEYWORDS_NO):
                return i
    col_lens: dict = {}
    for row in rows_sample:
        for i, txt in enumerate(row):
            if i == 0 or DATE_RE.match(txt) or not txt:
                continue
            col_lens.setdefault(i, []).append(len(txt))
    best, best_avg = 1, 0
    for ci, lens in col_lens.items():
        avg = sum(lens) / len(lens)
        if avg > best_avg:
            best_avg, best = avg, ci
    return best


def _find_col(headers: list, keywords: list) -> int | None:
    for kw in keywords:
        for i, h in enumerate(headers):
            if kw in h.upper():
                return i
    return None


# ── Selección de tabla ─────────────────────────────────────────────────────────

def elegir_tabla(soup: BeautifulSoup):
    tables = soup.find_all("table")
    if not tables:
        return None
    if len(tables) == 1:
        return tables[0]
    scored = []
    for t in tables:
        tbody = t.find("tbody") or t
        score = sum(
            1 for tr in tbody.find_all("tr")
            if tr.find_all("td") and re.sub(r'\D', '', tr.find_all("td")[0].get_text())
        )
        scored.append((score, len(tbody.find_all("tr")), t))
    scored.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return scored[0][2]


# ── Parser de tabla HTML ───────────────────────────────────────────────────────

def parse_tabla(soup: BeautifulSoup, page_url: str,
                seccion: str, pdf_base: str = "") -> list:
    table = elegir_tabla(soup)
    if not table:
        print("  [debug] No se encontró <table> en el HTML")
        return []

    headers = []
    thead = table.find("thead")
    if thead:
        for c in thead.find_all(["th", "td"]):
            headers.append(c.get_text(" ", strip=True))
    else:
        first_tr = table.find("tr")
        if first_tr:
            for c in first_tr.find_all(["th", "td"]):
                headers.append(c.get_text(" ", strip=True))

    tbody  = table.find("tbody") or table
    all_tr = tbody.find_all("tr")

    raw_rows = []
    seen: set = set()

    for tr in all_tr:
        cells = tr.find_all("td")
        if len(cells) < 3:
            continue
        num_txt = re.sub(r'\D', '', cells[0].get_text())
        if not num_txt:
            continue
        num = int(num_txt)
        if num in seen:
            continue
        seen.add(num)

        cell_texts = [c.get_text(" ", strip=True).split("\n")[0].strip() for c in cells]
        links = []
        for a in tr.find_all("a", href=True):
            img     = a.find("img")
            img_src = img.get("src", "") if img else ""
            links.append({"href": a["href"], "imgSrc": img_src})

        raw_rows.append((num, cell_texts, links))

    if not raw_rows:
        print(f"  [debug] {len(all_tr)} <tr> pero 0 pasaron el filtro de número")
        for i, tr in enumerate(all_tr[:5]):
            cells = tr.find_all("td")
            print(f"  [debug] tr[{i}] {len(cells)} celdas: "
                  f"{[c.get_text(' ', strip=True)[:25] for c in cells]}")
        return []

    # Detectar columna de título
    if seccion == "federal":
        col_titulo = 1
        col_pub = col_reform = col_tipo = None
    else:
        samples    = [r[1] for r in raw_rows[:12]]
        col_titulo = detectar_col_titulo(headers, samples)
        col_label  = headers[col_titulo] if col_titulo < len(headers) else "?"
        print(f"  → Headers   : {headers}")
        print(f"  → Col titulo: {col_titulo} → '{col_label}'")

        col_pub    = _find_col(headers, ["PUBLICACION", "PUBLICACIÓN", "FECHA PUB"])
        col_reform = _find_col(headers, ["REFORMA", "REFORMADO", "ÚLT. REFORMA", "ULT. REFORMA"])
        col_tipo   = _find_col(headers, ["TIPO"])

        if col_pub    is not None: print(f"  → Col pub   : {col_pub} → '{headers[col_pub]}'")
        if col_reform is not None: print(f"  → Col reform: {col_reform} → '{headers[col_reform]}'")
        if col_tipo   is not None: print(f"  → Col tipo  : {col_tipo} → '{headers[col_tipo]}'")

    for num, cells, links in raw_rows[:2]:
        print(f"    [debug] #{num:03d} | cells={cells} | links={len(links)}")

    docs = []
    for num, cell_texts, links in raw_rows:
        if col_titulo >= len(cell_texts):
            continue
        titulo = cell_texts[col_titulo]
        if not titulo:
            continue

        pdf_primary = pdf_alt = ""
        for lnk in links:
            if not es_pdf(lnk["href"], lnk["imgSrc"]):
                continue
            pdf_primary, pdf_alt = preparar_urls(lnk["href"], page_url, pdf_base)
            break

        if pdf_primary:
            doc = {"num": num, "titulo": titulo, "url": pdf_primary, "url_alt": pdf_alt}
            if seccion != "federal":
                doc["fecha_pub"]    = cell_texts[col_pub]    if col_pub    is not None and col_pub    < len(cell_texts) else ""
                doc["fecha_reforma"] = cell_texts[col_reform] if col_reform is not None and col_reform < len(cell_texts) else ""
                doc["tipo"]         = cell_texts[col_tipo]   if col_tipo   is not None and col_tipo   < len(cell_texts) else ""
            docs.append(doc)
        else:
            print(f"    ⚠ #{num:03d} '{titulo[:55]}' — sin enlace PDF")

    return docs


# ── Obtener HTML ───────────────────────────────────────────────────────────────

def get_html_requests(url: str, session: requests.Session) -> str | None:
    try:
        r = session.get(url, timeout=30, verify=False)
        r.raise_for_status()
        r.encoding = r.apparent_encoding or "utf-8"
        return r.text
    except Exception as e:
        print(f"  ⚠ requests falló: {e}")
        return None


def get_html_selenium(url: str) -> str | None:
    try:
        from selenium import webdriver
        from selenium.webdriver.chrome.options import Options
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait, Select
        from selenium.common.exceptions import NoSuchElementException, TimeoutException
    except ImportError:
        print("  ✗ selenium no instalado")
        return None

    opts = Options()
    opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(f"--user-agent={UA}")
    opts.add_argument("--ignore-certificate-errors")

    try:
        from webdriver_manager.chrome import ChromeDriverManager
        from selenium.webdriver.chrome.service import Service
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    except Exception:
        try:
            driver = webdriver.Chrome(options=opts)
        except Exception as e:
            print(f"  ✗ No se pudo iniciar Chrome: {e}")
            return None

    try:
        driver.get(url)
        try:
            WebDriverWait(driver, 35).until(
                lambda d: len(d.find_elements(By.CSS_SELECTOR, "table tr")) >= 3)
        except TimeoutException:
            return None

        time.sleep(2)

        for css in [".dataTables_length select", "select[name*='DataTables']", "select[name*='_length']"]:
            try:
                el = driver.find_element(By.CSS_SELECTOR, css)
                s  = Select(el)
                for op in s.options:
                    val = op.get_attribute("value") or ""
                    if val in ("-1", "All", "all") or "all" in op.text.lower():
                        s.select_by_value(val)
                        time.sleep(4)
                        break
                else:
                    s.select_by_index(len(s.options) - 1)
                    time.sleep(3)
            except NoSuchElementException:
                continue

        try:
            driver.execute_script(
                "if(typeof $!=='undefined'&&$.fn&&$.fn.DataTable){"
                "var t=$.fn.dataTable.tables(true);"
                "if(t&&t.length)$(t).DataTable().page.len(-1).draw();}"
            )
            time.sleep(4)
        except Exception:
            pass

        prev = 0
        for _ in range(10):
            n = len(driver.find_elements(By.CSS_SELECTOR, "table tbody tr"))
            if n > 0 and n == prev:
                break
            prev = n
            time.sleep(1)

        print(f"  → Selenium: {prev} filas en tbody")
        return driver.page_source
    finally:
        driver.quit()


# ── Scraper principal ──────────────────────────────────────────────────────────

_oaxaca_warmup_done = False


def _scrape_federal(session: requests.Session) -> list:
    """Parser especializado para diputados.gob.mx (encoding ISO-8859-1, HTML legacy)."""
    url = "https://www.diputados.gob.mx/LeyesBiblio/index.htm"
    pdf_base = "https://www.diputados.gob.mx/LeyesBiblio/"
    print(f"  Cargando Federal: {url}")

    try:
        r = session.get(url, timeout=60, verify=False)
        r.raise_for_status()
        html = r.content.decode('latin-1').replace('\r\n', '\n')
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return []

    # Parsear por bloques: cada fila empieza con un número de 3 dígitos
    positions = [(m.start(), m.group(1)) for m in
                 re.finditer(r'<font color="#595843" face="Verdana" size="2">(\d{3})</font>', html)]

    if not positions:
        print("  ✗ No se encontraron filas numeradas")
        return []

    docs = []
    for i, (pos, num_str) in enumerate(positions):
        end = positions[i + 1][0] if i + 1 < len(positions) else len(html)
        block = html[pos:end]
        num = int(num_str)

        # Título: texto dentro del <a href="ref/..."> (columna 2)
        m_title = re.search(
            r'<a\s+href="ref/[^"]+\.htm"[^>]*>\s*<font[^>]*>(.*?)</font>\s*</a>',
            block, re.DOTALL
        )
        if m_title:
            titulo = re.sub(r'<[^>]+>', '', m_title.group(1)).strip()
            titulo = re.sub(r'\s+', ' ', titulo)
        else:
            m_bold = re.search(r'<b>\s*<font[^>]*>(.*?)</font>', block, re.DOTALL)
            titulo = re.sub(r'<[^>]+>', '', m_bold.group(1)).strip() if m_bold else ""
            titulo = re.sub(r'\s+', ' ', titulo)

        if not titulo:
            continue

        # PDF: href="pdf/XXX.pdf"
        m_pdf = re.search(r'href="(pdf/[^"]+\.pdf)"', block)
        if not m_pdf:
            print(f"    ⚠ #{num:03d} '{titulo[:55]}' — sin enlace PDF")
            continue
        url_pdf = pdf_base + m_pdf.group(1)

        # Fechas DOF
        dof_dates = re.findall(r'DOF\s+(\d{2}/\d{2}/\d{4})', block)
        fecha_pub = dof_dates[0] if dof_dates else ""
        fecha_reforma = dof_dates[1] if len(dof_dates) > 1 else ""

        docs.append({
            "num": num,
            "titulo": titulo,
            "url": url_pdf,
            "url_alt": "",
            "fecha_pub": fecha_pub,
            "fecha_reforma": fecha_reforma,
        })

    print(f"  → {len(docs)} docs encontrados (regex/ISO-8859-1)")
    return docs


def scrape_seccion(nombre: str, cfg: dict, session: requests.Session) -> list:
    global _oaxaca_warmup_done
    url      = cfg["url"]
    pdf_base = cfg.get("pdf_base", "")
    tipo     = cfg["tipo"]

    # Federal usa parser especializado (HTML legacy con encoding ISO-8859-1)
    if tipo == "federal":
        return _scrape_federal(session)

    if tipo == "oaxaca" and not _oaxaca_warmup_done:
        try:
            session.get("https://www.congresooaxaca.gob.mx/", timeout=10, verify=False)
            _oaxaca_warmup_done = True
        except Exception:
            pass

    print(f"  Cargando (requests): {url}")
    html = get_html_requests(url, session)

    if html:
        for parser in ("lxml", "html.parser"):
            try:
                soup = BeautifulSoup(html, parser)
                docs = parse_tabla(soup, url, tipo, pdf_base)
                if docs:
                    print(f"  → {len(docs)} docs encontrados (requests/{parser})")
                    return docs
            except Exception as e:
                print(f"  ⚠ {parser}: {e}")
        print("  ⚠ requests: 0 filas con PDF. Intentando Selenium…")

    print(f"  Cargando (Selenium): {url}")
    html = get_html_selenium(url)
    if not html:
        print("  ✗ Selenium también falló.")
        return []

    for parser in ("lxml", "html.parser"):
        try:
            soup = BeautifulSoup(html, parser)
            docs = parse_tabla(soup, url, tipo, pdf_base)
            if docs:
                print(f"  → {len(docs)} docs encontrados (Selenium/{parser})")
                return docs
        except Exception as e:
            print(f"  ⚠ {parser}: {e}")

    print("  ✗ 0 documentos encontrados.")
    return []


# ── Descarga concurrente ───────────────────────────────────────────────────────

def _descargar_pdf(urls: list, dest: Path,
                   session: requests.Session, referer: str) -> tuple[bool, str]:
    dl_headers = {
        **HEADERS_BASE,
        "Referer": referer,
        "Accept":  "application/pdf,application/octet-stream,*/*",
    }
    last_err = "Sin URLs"
    for url in urls:
        if not url:
            continue
        try:
            r = session.get(url, headers=dl_headers, timeout=DL_TIMEOUT, stream=True, verify=False)
            r.raise_for_status()
            ct = r.headers.get("Content-Type", "")
            if "text/html" in ct.lower():
                last_err = f"HTML devuelto ({url[:60]}…)"
                continue
            dest.parent.mkdir(parents=True, exist_ok=True)
            with open(dest, "wb") as f:
                for chunk in r.iter_content(65_536):
                    if chunk:
                        f.write(chunk)
            if dest.stat().st_size < 1024:
                dest.unlink(missing_ok=True)
                last_err = f"Archivo vacío desde {url[:60]}…"
                continue
            return True, url
        except requests.exceptions.HTTPError as e:
            last_err = f"HTTP {e.response.status_code} ({url[:60]}…)"
        except requests.exceptions.ConnectionError:
            last_err = f"Error de conexión ({url[:60]}…)"
        except Exception as e:
            last_err = f"{type(e).__name__}: {str(e)[:60]}"
    return False, last_err


def descargar_seccion(nombre: str, docs: list, carpeta: Path,
                      session: requests.Session):
    carpeta.mkdir(parents=True, exist_ok=True)
    referer  = next((cfg["url"] for n, cfg in SECCIONES.items() if n == nombre),
                    "https://www.congresooaxaca.gob.mx/")
    tareas   = []
    omitidos = 0

    for d in docs:
        base = aplicar_nombre(d["num"], d["titulo"], nombre)
        dest = carpeta / f"{base}.pdf"
        # Skip si ya existe PDF con ese nombre o número
        existing = _any_file(carpeta, d["num"], "pdf", nombre_base=base)
        if existing and existing.stat().st_size > 1024:
            omitidos += 1
            continue
        urls = [u for u in [d.get("url", ""), d.get("url_alt", "")] if u]
        tareas.append((urls, dest, base))

    if omitidos:
        print(f"  ↩  {omitidos} ya descargados — omitidos")
    if not tareas:
        print("  ✓  Sin descargas nuevas.")
    else:
        print(f"  → Descargando {len(tareas)} PDFs ({MAX_WORKERS} hilos)\n")
        ok = fail = 0
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
            futuros = {
                ex.submit(_descargar_pdf, urls, dest, session, referer): base
                for urls, dest, base in tareas
            }
            for fut in as_completed(futuros):
                base = futuros[fut]
                try:
                    exito, info = fut.result()
                except Exception as e:
                    exito, info = False, str(e)
                if exito:
                    ok += 1
                    print(f"  ✓  {base}.pdf")
                else:
                    fail += 1
                    print(f"  ✗  {base}.pdf  ← {info}")

        nums_esperados = {d["num"] for d in docs}
        pdfs_ok = [f for f in carpeta.iterdir()
                   if f.suffix.lower() == ".pdf" and f.stat().st_size > 1024]
        mapa: dict = {}
        for f in pdfs_ok:
            m = re.match(r'^(\d+)\s*-', f.name)
            if m:
                mapa.setdefault(int(m.group(1)), []).append(f.name)

        presentes = set(mapa.keys())
        faltantes = sorted(nums_esperados - presentes)
        dupls     = {n: fs for n, fs in mapa.items() if len(fs) > 1}
        extra     = sorted(presentes - nums_esperados)

        print(f"\n  {'─'*58}")
        print(f"  Validación [{nombre}]")
        print(f"  Esperados:{len(nums_esperados)} | Disco:{len(pdfs_ok)} | "
              f"OK:{ok} | Fallo:{fail} | Omitido:{omitidos}")
        if faltantes:
            print(f"  ✗ Faltantes ({len(faltantes)}): {faltantes}")
        else:
            print(f"  ✓ Sin faltantes — todos presentes")
        if dupls:
            for n, fs in sorted(dupls.items()):
                print(f"  ⚠ Duplicado #{n:03d}: {fs}")
        else:
            print(f"  ✓ Sin duplicados")
        if extra:
            print(f"  ⚠ Números extra: {extra}")
        print(f"  {'─'*58}\n")


# ── Main ───────────────────────────────────────────────────────────────────────

# ############################################################
#  COMANDO: FIX (agregar campos faltantes a JSONs existentes)
# ############################################################

def _fix_calcular_calidad(contenido, numero_paginas, metodo):
    if not contenido or not contenido.strip():
        return {"calidad_porcentaje": 0, "calidad_detalle": "Sin contenido"}
    chars = len(contenido)
    paginas = max(numero_paginas, 1)
    score_densidad = min((chars / paginas) / 2000, 1.0)
    alfanum = sum(1 for c in contenido if c.isalnum() or c in ' \n\t.,;:()áéíóúñüÁÉÍÓÚÑÜ"\'—–-')
    ratio_alfanum = alfanum / max(chars, 1)
    lineas = contenido.split('\n')
    lineas_con_texto = sum(1 for l in lineas if len(l.strip()) > 5)
    ratio_lineas = lineas_con_texto / max(len(lineas), 1)
    palabras = contenido.split()
    palabras_cortas = sum(1 for p in palabras if len(p) == 1 and p.lower() not in 'aeiouy')
    ratio_basura = 1.0 - min(palabras_cortas / max(len(palabras), 1) * 5, 0.5)
    calidad = score_densidad * 0.30 + ratio_alfanum * 0.25 + ratio_lineas * 0.20 + ratio_basura * 0.25
    porcentaje = round(min(calidad * 100, 100))
    if porcentaje >= 85:   detalle = "Excelente"
    elif porcentaje >= 70: detalle = "Buena"
    elif porcentaje >= 50: detalle = "Aceptable"
    elif porcentaje >= 30: detalle = "Baja"
    else:                  detalle = "Muy baja"
    return {"calidad_porcentaje": porcentaje, "calidad_detalle": detalle}


def _fix_extraer_fecha_reforma(titulo, contenido):
    meses = {
        'enero': '01', 'febrero': '02', 'marzo': '03', 'abril': '04',
        'mayo': '05', 'junio': '06', 'julio': '07', 'agosto': '08',
        'septiembre': '09', 'octubre': '10', 'noviembre': '11', 'diciembre': '12'
    }
    fuentes = []
    if contenido:
        fuentes.append(contenido[:15000].lower())
    if titulo:
        fuentes.append(titulo.lower())
    for fuente in fuentes:
        patron = r'[uú]ltima\s+reforma\s*:?\s*[^\n]{0,120}?(\d{1,2})[°º]?\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de(?:l)?\s+(\d{4})'
        match = re.search(patron, fuente)
        if match:
            return f"{match.group(1).zfill(2)}/{meses[match.group(2)]}/{match.group(3)}"
    if not contenido:
        return ""
    patron_reforma = r'(?:[uú]ltima\s+)?reform[aáo][^.]{0,80}?(\d{1,2})\s+de\s+(enero|febrero|marzo|abril|mayo|junio|julio|agosto|septiembre|octubre|noviembre|diciembre)\s+de(?:l)?\s+(\d{4})'
    matches = re.findall(patron_reforma, contenido.lower()[:20000])
    if matches:
        fechas = []
        for m in matches:
            try:
                dia, anio = int(m[0]), int(m[2])
                mes_num = int(meses.get(m[1], 1))
                if 1900 < anio < 2100 and 1 <= dia <= 31:
                    fechas.append((anio, mes_num, dia, f"{str(dia).zfill(2)}/{meses[m[1]]}/{m[2]}"))
            except (ValueError, KeyError):
                continue
        if fechas:
            fechas.sort(reverse=True)
            return fechas[0][3]
    return ""


def _fix_detectar_metodo(contenido, es_escaneado):
    if not contenido or len(contenido.strip()) < 50:
        return "ninguno"
    if "[Pág" in contenido and "OCR]" in contenido:
        return "pdfplumber+paddle"
    if es_escaneado:
        return "pdfplumber+paddle"
    return "pdfplumber"


def cmd_fix():
    """Agrega campos v8 faltantes a los JSONs existentes."""
    import glob as _glob
    print("=" * 66)
    print("  FIX JSONs OAXACA — Agregar campos faltantes")
    print("=" * 66)

    secciones = ["Estatal", "Federal", "Marco Normativo", "Municipal"]
    total_fixed = 0
    total_fecha = 0
    total_files = 0

    for seccion in secciones:
        json_dir = BASE_DIR_TEXTO / seccion
        if not json_dir.exists():
            continue
        archivos = sorted(_glob.glob(str(json_dir / "*.json")))
        fixed_sec = fecha_sec = 0
        print(f"\n  {seccion}: {len(archivos)} archivos")

        for ruta in archivos:
            total_files += 1
            try:
                with open(ruta, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            except Exception:
                continue

            modificado = False
            contenido = data.get("contenido", "")
            titulo = data.get("titulo", "")
            num_pags = data.get("numero_paginas", 0) or 0
            es_escaneado = data.get("es_escaneado", False)

            if "metodo_extraccion" not in data:
                data["metodo_extraccion"] = _fix_detectar_metodo(contenido, es_escaneado)
                modificado = True
            if "calidad_porcentaje" not in data or "calidad_detalle" not in data:
                cal = _fix_calcular_calidad(contenido, num_pags, data.get("metodo_extraccion", ""))
                data["calidad_porcentaje"] = cal["calidad_porcentaje"]
                data["calidad_detalle"] = cal["calidad_detalle"]
                modificado = True
            if not data.get("fecha_ultima_reforma"):
                fecha = _fix_extraer_fecha_reforma(titulo, contenido)
                if fecha:
                    data["fecha_ultima_reforma"] = fecha
                    fecha_sec += 1
                    modificado = True

            if modificado:
                with open(ruta, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                fixed_sec += 1

        total_fixed += fixed_sec
        total_fecha += fecha_sec
        print(f"    Campos agregados: {fixed_sec}")
        print(f"    Fechas reforma recuperadas: {fecha_sec}")

    print(f"\n{'='*66}")
    print(f"  TOTAL: {total_fixed}/{total_files} JSONs actualizados")
    print(f"  Fechas reforma recuperadas: {total_fecha}")
    print(f"{'='*66}")


# ############################################################
#  COMANDO: VALIDAR (verificar estructura de JSONs)
# ############################################################

_VAL_CAMPOS = {
    "titulo":             str,
    "tipo_ordenamiento":  str,
    "jurisdiccion":       str,
    "materia":            list,
    "fecha_publicacion":  (str, type(None)),
    "es_vigente":         (bool, type(None)),
    "status_texto":       (str, type(None)),
    "contenido":          str,
    "url_fuente":         (str, type(None)),
    "es_escaneado":       bool,
    "calidad_ocr":        (int, float, type(None)),
    "archivo_origen":     str,
}
_VAL_TIPOS = ["Ley", "Código", "Reglamento", "Decreto", "Constitución",
              "Acuerdo", "NOM", "Estatuto", "Ordenanza", "Presupuesto", "Otro"]
_VAL_STATUS = ["Vigente", "Abrogada", "Derogada", None]
_VAL_FECHA_RE = re.compile(r'^\d{4}-\d{2}-\d{2}$')


def cmd_validar(seccion_nombre="Federal"):
    """Valida los JSONs de una sección contra la estructura esperada."""
    from datetime import datetime as _dt

    json_dir = str(BASE_DIR_TEXTO / seccion_nombre)
    pdf_dir = str(BASE_DIR / seccion_nombre)

    if not Path(json_dir).exists():
        print(f"  Sección '{seccion_nombre}' no encontrada en {json_dir}")
        return

    print("=" * 70)
    print(f"  VALIDADOR — SECCIÓN {seccion_nombre.upper()} (OAXACA)")
    print(f"  {_dt.now().strftime('%d/%m/%Y %H:%M')}")
    print("=" * 70)

    jsons = sorted([f for f in Path(json_dir).iterdir() if f.suffix == '.json'])
    pdfs = set(f.name for f in Path(pdf_dir).iterdir() if f.suffix.lower() == '.pdf') if Path(pdf_dir).exists() else set()

    print(f"\n  JSONs: {len(jsons)} | PDFs: {len(pdfs)}")

    total = len(jsons)
    ok = 0
    problemas_log = []
    materias_g = {}
    tipos_g = {}
    vigentes = no_vigentes = nulos_v = 0
    sin_url = sin_contenido = sin_pdf = 0

    for jpath in jsons:
        jf = jpath.name
        problemas = []
        try:
            with open(jpath, 'r', encoding='utf-8') as f:
                d = json.load(f)
        except Exception as e:
            problemas_log.append((jf, f"JSON inválido: {e}"))
            continue

        # Campos y tipos
        for campo, tipo_esp in _VAL_CAMPOS.items():
            if campo not in d:
                problemas.append(f"Falta: {campo}")
            else:
                val = d[campo]
                if isinstance(tipo_esp, tuple):
                    if not isinstance(val, tipo_esp):
                        problemas.append(f"{campo}: tipo {type(val).__name__}")
                elif not isinstance(val, tipo_esp):
                    problemas.append(f"{campo}: tipo {type(val).__name__}")

        # Titulo = nombre archivo
        titulo = d.get('titulo', '')
        if titulo.lower().strip() != jf.replace('.json', '').lower().strip():
            problemas.append("titulo ≠ nombre archivo")

        # tipo_ordenamiento
        tipo_ord = d.get('tipo_ordenamiento', '')
        if tipo_ord not in _VAL_TIPOS:
            problemas.append(f"tipo_ordenamiento: '{tipo_ord}'")
        tipos_g[tipo_ord] = tipos_g.get(tipo_ord, 0) + 1

        # jurisdiccion
        if d.get('jurisdiccion') != 'Oaxaca':
            problemas.append(f"jurisdiccion: '{d.get('jurisdiccion')}'")

        # materia
        materia = d.get('materia', [])
        if not isinstance(materia, list) or not materia:
            problemas.append("materia vacía")
        else:
            for m in materia:
                materias_g[m] = materias_g.get(m, 0) + 1

        # fecha ISO
        fecha = d.get('fecha_publicacion')
        if fecha is not None and not _VAL_FECHA_RE.match(str(fecha)):
            problemas.append(f"fecha formato: '{fecha}'")

        # vigencia
        es_vig = d.get('es_vigente')
        status = d.get('status_texto')
        if es_vig is True:
            vigentes += 1
        elif es_vig is False:
            no_vigentes += 1
        else:
            nulos_v += 1
        if es_vig is True and status not in ('Vigente', None):
            problemas.append(f"vigente/status incoherente")
        if es_vig is False and status not in ('Abrogada', 'Derogada', None):
            problemas.append(f"no vigente/status incoherente")

        # contenido
        if len(d.get('contenido', '')) < 100:
            sin_contenido += 1
            problemas.append("contenido corto")

        # url
        if not d.get('url_fuente'):
            sin_url += 1

        # archivo_origen
        archivo = d.get('archivo_origen', '')
        if archivo not in pdfs:
            sin_pdf += 1
            problemas.append("archivo_origen sin PDF")

        if not problemas:
            ok += 1
        elif len(problemas_log) < 20:
            for p in problemas:
                problemas_log.append((jf, p))

    # Reporte
    print(f"\n{'─'*70}")
    print(f"  ✓ Sin problemas: {ok}/{total}")
    print(f"  ⚠ Con problemas: {total - ok}")
    print(f"  Vigentes: {vigentes} | No vigentes: {no_vigentes} | Null: {nulos_v}")
    print(f"  Sin contenido: {sin_contenido} | Sin URL: {sin_url} | Sin PDF: {sin_pdf}")

    print(f"\n  Tipos de ordenamiento:")
    for t, n in sorted(tipos_g.items(), key=lambda x: -x[1]):
        print(f"    {t:<20} {n:>4}")

    print(f"\n  Materias ({len(materias_g)}):")
    for m, n in sorted(materias_g.items(), key=lambda x: -x[1])[:15]:
        barra = '█' * min(n, 40)
        print(f"    {m:<25} {n:>4}  {barra}")

    if problemas_log:
        print(f"\n  Problemas (primeros 20):")
        for jf, p in problemas_log[:20]:
            print(f"    {jf[:40]} → {p}")

    # PDFs sin JSON
    j_names = set(j.stem for j in jsons)
    p_names = set(p.replace('.pdf', '') for p in pdfs)
    sin_j = p_names - j_names
    if sin_j:
        print(f"\n  PDFs sin JSON ({len(sin_j)}):")
        for p in sorted(sin_j)[:5]:
            print(f"    {p[:60]}.pdf")

    print(f"\n{'='*70}")
    if ok == total:
        print(f"  ✓ VALIDACIÓN EXITOSA — {total} JSONs correctos")
    else:
        print(f"  ⚠ {ok}/{total} correctos")
    print(f"{'='*70}")


# ############################################################
#  COMANDO: REPORTE (generar Excel)
# ############################################################

def cmd_reporte():
    """Genera reporte Excel con resumen y detalle por sección."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl no instalado. pip install openpyxl")
        return
    from datetime import datetime as _dt

    print("=" * 66)
    print("  REPORTE EXCEL — OAXACA")
    print("=" * 66)

    secciones = ["Estatal", "Federal", "Marco Normativo", "Municipal"]
    wb = Workbook()
    hf = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    alt = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    f_exc = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    f_bue = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    f_baj = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    brd = Border(left=Side(style='thin', color='D9E2F3'), right=Side(style='thin', color='D9E2F3'),
                 top=Side(style='thin', color='D9E2F3'), bottom=Side(style='thin', color='D9E2F3'))

    ws = wb.active
    ws.title = "Resumen General"
    ws.merge_cells('A1:I1')
    ws['A1'] = "Reporte de Legislación — Oaxaca"
    ws['A1'].font = Font(name='Calibri', bold=True, size=14, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:I2')
    ws['A2'] = f"Generado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(name='Calibri', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')

    hdrs = ['Sección', 'JSONs', 'PDFs', 'Con Texto', 'Sin Texto', 'Calidad Prom.', 'URLs Dup.', 'Con Fecha Reforma', 'Páginas Total']
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = brd

    row = 5
    for seccion in secciones:
        jd = BASE_DIR_TEXTO / seccion
        pd = BASE_DIR / seccion
        nj = np_ = ct = st = pgs = cf = 0
        cals = []; urls = set(); dup = 0
        if jd.exists():
            for jf in jd.iterdir():
                if jf.suffix != '.json': continue
                nj += 1
                try:
                    d = json.loads(jf.read_text(encoding='utf-8'))
                    c_ = d.get('contenido', '')
                    u = d.get('url', d.get('url_fuente', ''))
                    cals.append(d.get('calidad_porcentaje', d.get('calidad_ocr', 0)) or 0)
                    pgs += d.get('numero_paginas', 0) or 0
                    if c_ and len(c_) > 100: ct += 1
                    else: st += 1
                    if d.get('fecha_ultima_reforma') or d.get('fecha_publicacion'): cf += 1
                    if u in urls: dup += 1
                    urls.add(u)
                except: pass
        if pd.exists():
            np_ = sum(1 for f in pd.iterdir() if f.suffix.lower() == '.pdf')
        cal = sum(cals) / max(len(cals), 1)
        ws.cell(row=row, column=1, value=seccion).border = brd
        for c, v in enumerate([nj, np_, ct, st, f"{cal:.0f}%", dup, cf, pgs], 2):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = brd; cell.alignment = Alignment(horizontal='center')
        cc = ws.cell(row=row, column=6)
        if cal >= 85: cc.fill = f_exc
        elif cal >= 70: cc.fill = f_bue
        elif cal > 0: cc.fill = f_baj
        row += 1

    ws.column_dimensions['A'].width = 22
    for c in range(2, 10): ws.column_dimensions[get_column_letter(c)].width = 15
    ws.freeze_panes = 'A5'

    # Hojas detalle
    dh = ['#', 'Título', 'Tipo', 'Status', 'Materia', 'Método', 'Págs', 'Calidad %', 'Calidad', 'Fecha Pub.', 'Fecha Reforma', 'Escaneado', 'Chars', 'URL']
    for seccion in secciones:
        jd = BASE_DIR_TEXTO / seccion
        if not jd.exists(): continue
        wss = wb.create_sheet(title=seccion[:31])
        for c, h in enumerate(dh, 1):
            cell = wss.cell(row=1, column=c, value=h)
            cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = brd
        r = 2
        for jf in sorted(jd.iterdir()):
            if jf.suffix != '.json': continue
            try: d = json.loads(jf.read_text(encoding='utf-8'))
            except: continue
            nm = re.match(r'^(\d+)', jf.name)
            mat = d.get('materia', [])
            if isinstance(mat, list): mat = ', '.join(mat)
            wss.cell(row=r, column=1, value=int(nm.group(1)) if nm else r-1).border = brd
            wss.cell(row=r, column=2, value=d.get('titulo', '')[:200]).border = brd
            wss.cell(row=r, column=3, value=d.get('ordenamiento', d.get('tipo_ordenamiento', ''))).border = brd
            wss.cell(row=r, column=4, value=d.get('status', d.get('status_texto', ''))).border = brd
            wss.cell(row=r, column=5, value=mat).border = brd
            wss.cell(row=r, column=6, value=d.get('metodo_extraccion', '')).border = brd
            wss.cell(row=r, column=7, value=d.get('numero_paginas', 0)).border = brd
            cal = d.get('calidad_porcentaje', d.get('calidad_ocr', 0)) or 0
            cc = wss.cell(row=r, column=8, value=cal); cc.border = brd
            if cal >= 85: cc.fill = f_exc
            elif cal >= 70: cc.fill = f_bue
            elif cal > 0: cc.fill = f_baj
            wss.cell(row=r, column=9, value=d.get('calidad_detalle', '')).border = brd
            wss.cell(row=r, column=10, value=d.get('fecha_publicacion', '')).border = brd
            wss.cell(row=r, column=11, value=d.get('fecha_ultima_reforma', '')).border = brd
            wss.cell(row=r, column=12, value='Sí' if d.get('es_escaneado') else 'No').border = brd
            wss.cell(row=r, column=13, value=len(d.get('contenido', ''))).border = brd
            wss.cell(row=r, column=14, value=d.get('url', d.get('url_fuente', ''))).border = brd
            if (r-2)%2==1:
                for c in range(1,15):
                    if wss.cell(row=r,column=c).fill==PatternFill(): wss.cell(row=r,column=c).fill=alt
            r += 1
        wss.column_dimensions['A'].width = 5; wss.column_dimensions['B'].width = 65
        wss.column_dimensions['C'].width = 15; wss.column_dimensions['N'].width = 50
        wss.freeze_panes = 'A2'; wss.auto_filter.ref = f"A1:N{r-1}"

    # ── HOJA: No Procesados ──
    ws_err = wb.create_sheet(title="No Procesados")
    err_headers = ['Sección', 'Archivo', 'Problema', 'Detalle']
    for c, h in enumerate(err_headers, 1):
        cell = ws_err.cell(row=1, column=c, value=h)
        cell.font = hf; cell.fill = hfill; cell.alignment = Alignment(horizontal='center'); cell.border = brd

    err_row = 2
    fill_err = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for seccion in secciones:
        jd = BASE_DIR_TEXTO / seccion
        pd = BASE_DIR / seccion
        if not jd.exists():
            continue

        json_names = set(f.stem for f in jd.iterdir() if f.suffix == '.json')
        pdf_names = set(f.stem for f in pd.iterdir() if f.suffix.lower() == '.pdf') if pd.exists() else set()

        # JSONs sin PDF
        for nombre in sorted(json_names - pdf_names):
            for c, v in enumerate([seccion, f"{nombre}.json", "JSON sin PDF correspondiente", "El PDF no existe o fue renombrado"], 1):
                cell = ws_err.cell(row=err_row, column=c, value=v)
                cell.border = brd
                cell.fill = fill_err if c == 3 else PatternFill()
            err_row += 1

        # PDFs sin JSON
        for nombre in sorted(pdf_names - json_names):
            for c, v in enumerate([seccion, f"{nombre}.pdf", "PDF sin JSON correspondiente", "No se pudo extraer texto"], 1):
                cell = ws_err.cell(row=err_row, column=c, value=v)
                cell.border = brd
                cell.fill = fill_err if c == 3 else PatternFill()
            err_row += 1

        # JSONs con contenido vacío/corto
        for jf in sorted(jd.iterdir()):
            if jf.suffix != '.json':
                continue
            try:
                d = json.loads(jf.read_text(encoding='utf-8'))
                contenido = d.get('contenido', '')
                if len(contenido) < 100:
                    # No duplicar si ya reportamos como "sin PDF"
                    if jf.stem in json_names & pdf_names or jf.stem not in (json_names - pdf_names):
                        for c, v in enumerate([seccion, jf.name, "Contenido vacío o muy corto", f"{len(contenido)} caracteres extraídos"], 1):
                            cell = ws_err.cell(row=err_row, column=c, value=v)
                            cell.border = brd
                            cell.fill = fill_err if c == 3 else PatternFill()
                        err_row += 1
            except:
                for c, v in enumerate([seccion, jf.name, "JSON corrupto", "No se pudo leer el archivo"], 1):
                    cell = ws_err.cell(row=err_row, column=c, value=v)
                    cell.border = brd
                    cell.fill = fill_err if c == 3 else PatternFill()
                err_row += 1

    ws_err.column_dimensions['A'].width = 18
    ws_err.column_dimensions['B'].width = 65
    ws_err.column_dimensions['C'].width = 30
    ws_err.column_dimensions['D'].width = 45
    ws_err.freeze_panes = 'A2'
    if err_row > 2:
        ws_err.auto_filter.ref = f"A1:D{err_row-1}"

    total_errores = err_row - 2
    print(f"  Documentos no procesados: {total_errores}")

    ruta = Path(BASE_DIR_TEXTO).parent / "reporte_oaxaca.xlsx"
    wb.save(str(ruta))
    print(f"  Reporte guardado: {ruta}")


# ############################################################
#  COMANDO: SCRAPE (scrape completo — original)
# ############################################################

def cmd_scrape():
    print("═" * 66)
    print("  SCRAPER v8 — Leyes de Oaxaca  (requests + BeautifulSoup + OCR)")
    print("═" * 66)
    if not _PDFPLUMBER_OK:
        print("  ⚠ pdfplumber no instalado")
    if _PADDLE_OK:
        print("  ✓ PaddleOCR disponible")
    elif _TESSERACT_OK:
        print("  ✓ Tesseract disponible")

    BASE_DIR.mkdir(exist_ok=True)
    BASE_DIR_TEXTO.mkdir(exist_ok=True)

    session = requests.Session()
    session.headers.update(HEADERS_BASE)

    for nombre, cfg in SECCIONES.items():
        print(f"\n{'═' * 66}")
        print(f"  SECCIÓN : {nombre}")
        print(f"  URL     : {cfg['url']}")
        print(f"{'─' * 66}")

        carpeta = BASE_DIR / nombre

        try:
            docs = scrape_seccion(nombre, cfg, session)
            docs.sort(key=lambda d: d["num"])

            if docs:
                descargar_seccion(nombre, docs, carpeta, session)
                print(f"\n  {'─'*66}")
                print(f"  EXTRACCIÓN DE TEXTO [{nombre}]")
                print(f"  {'─'*66}")
                procesar_textos_seccion(nombre, docs, carpeta)
            else:
                print("  ! Sin documentos PDF encontrados.")

        except Exception:
            import traceback
            print(f"  ✗ Error en [{nombre}]:")
            traceback.print_exc()

    print(f"\n{'═' * 66}")
    print(f"  COMPLETADO")
    print(f"  PDFs   → {BASE_DIR.resolve()}")
    print(f"  Textos → {BASE_DIR_TEXTO.resolve()}")
    print(f"{'═' * 66}")


# ############################################################
#  MAIN — CLI con subcomandos
# ############################################################

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='Scraper de Legislación — Oaxaca',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  %(prog)s scrape                Scrape completo (4 secciones)
  %(prog)s fix                   Agregar campos v8 faltantes
  %(prog)s validar               Validar JSONs de Federal
  %(prog)s validar Estatal       Validar JSONs de Estatal
  %(prog)s reporte               Generar reporte Excel
        """
    )
    sub = parser.add_subparsers(dest='comando')
    sub.add_parser('scrape', help='Scrape completo (4 secciones)')
    sub.add_parser('fix', help='Agregar campos v8 faltantes a JSONs existentes')

    p_val = sub.add_parser('validar', help='Validar estructura de JSONs')
    p_val.add_argument('seccion', nargs='?', default='Federal', help='Sección a validar (default: Federal)')

    sub.add_parser('reporte', help='Generar reporte Excel')

    args = parser.parse_args()

    if not args.comando:
        parser.print_help()
        return

    if args.comando == 'scrape':
        cmd_scrape()
    elif args.comando == 'fix':
        cmd_fix()
    elif args.comando == 'validar':
        cmd_validar(args.seccion)
    elif args.comando == 'reporte':
        cmd_reporte()


if __name__ == "__main__":
    main()
